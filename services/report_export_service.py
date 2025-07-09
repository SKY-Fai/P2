"""
Report Export Service for F-AI Accountant
Handles export of financial reports in multiple formats: Excel, Word, PDF
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import tempfile


class ReportExportService:
    """Service for exporting financial reports in multiple formats"""
    
    def __init__(self):
        self.reports_dir = 'reports'
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def export_report(self, report_data, report_name, format_type='excel', company_name="F-AI Accountant"):
        """
        Export report in specified format
        
        Args:
            report_data: Dictionary containing report data
            report_name: Name of the report
            format_type: 'excel', 'word', or 'pdf'
            company_name: Company name for headers
            
        Returns:
            str: Path to the generated file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format_type == 'excel':
            return self._export_to_excel(report_data, report_name, timestamp, company_name)
        elif format_type == 'word':
            return self._export_to_word(report_data, report_name, timestamp, company_name)
        elif format_type == 'pdf':
            return self._export_to_pdf(report_data, report_name, timestamp, company_name)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    def _export_to_excel(self, report_data, report_name, timestamp, company_name):
        """Export report to Excel format"""
        filename = f"{report_name}_{timestamp}.xlsx"
        file_path = os.path.join(self.reports_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_name
        
        # Professional styling
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(bold=True, size=16, color='1F4E79')
        company_font = Font(bold=True, size=14, color='666666')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Add company header
        ws.cell(row=current_row, column=1, value=company_name).font = company_font
        current_row += 1
        
        # Add report title
        ws.cell(row=current_row, column=1, value=report_name.replace('_', ' ').title()).font = title_font
        current_row += 1
        
        # Add generation date
        ws.cell(row=current_row, column=1, value=f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        current_row += 2
        
        # Process report data
        if isinstance(report_data, dict):
            for section_name, section_data in report_data.items():
                if isinstance(section_data, list) and section_data:
                    # Add section header
                    ws.cell(row=current_row, column=1, value=section_name.replace('_', ' ').title()).font = Font(bold=True, size=12)
                    current_row += 1
                    
                    # Add table headers
                    headers = list(section_data[0].keys()) if section_data else []
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header.replace('_', ' ').title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    current_row += 1
                    
                    # Add data rows
                    for row_data in section_data:
                        for col_idx, value in enumerate(row_data.values(), 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            cell.border = thin_border
                        current_row += 1
                    
                    current_row += 1  # Add space between sections
        
        elif isinstance(report_data, list) and report_data:
            # Single table format
            headers = list(report_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1
            
            for row_data in report_data:
                for col_idx, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        return file_path
    
    def _export_to_pdf(self, report_data, report_name, timestamp, company_name):
        """Export report to PDF format"""
        filename = f"{report_name}_{timestamp}.pdf"
        file_path = os.path.join(self.reports_dir, filename)
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#1F4E79'),
            alignment=1  # Center alignment
        )
        
        company_style = ParagraphStyle(
            'CompanyStyle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#666666'),
            alignment=1
        )
        
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.HexColor('#1F4E79')
        )
        
        story = []
        
        # Add company header
        story.append(Paragraph(company_name, company_style))
        story.append(Spacer(1, 12))
        
        # Add report title
        story.append(Paragraph(report_name.replace('_', ' ').title(), title_style))
        story.append(Spacer(1, 12))
        
        # Add generation date
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        story.append(Spacer(1, 24))
        
        # Process report data
        if isinstance(report_data, dict):
            for section_name, section_data in report_data.items():
                if isinstance(section_data, list) and section_data:
                    # Add section header
                    story.append(Paragraph(section_name.replace('_', ' ').title(), section_style))
                    story.append(Spacer(1, 12))
                    
                    # Create table
                    headers = list(section_data[0].keys()) if section_data else []
                    table_data = [[header.replace('_', ' ').title() for header in headers]]
                    
                    for row_data in section_data:
                        table_data.append([str(value) for value in row_data.values()])
                    
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(table)
                    story.append(Spacer(1, 24))
        
        elif isinstance(report_data, list) and report_data:
            # Single table format
            headers = list(report_data[0].keys())
            table_data = [[header.replace('_', ' ').title() for header in headers]]
            
            for row_data in report_data:
                table_data.append([str(value) for value in row_data.values()])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        doc.build(story)
        return file_path
    
    def _export_to_word(self, report_data, report_name, timestamp, company_name):
        """Export report to Word format (using HTML for simplicity)"""
        filename = f"{report_name}_{timestamp}.doc"
        file_path = os.path.join(self.reports_dir, filename)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .company-name {{ font-size: 18px; color: #666666; font-weight: bold; }}
                .report-title {{ font-size: 24px; color: #1F4E79; font-weight: bold; margin: 10px 0; }}
                .date {{ font-size: 12px; color: #666666; margin-bottom: 30px; }}
                .section-title {{ font-size: 16px; color: #1F4E79; font-weight: bold; margin: 20px 0 10px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
                th {{ background-color: #1F4E79; color: white; padding: 12px; text-align: center; font-weight: bold; }}
                td {{ padding: 8px 12px; border: 1px solid #ddd; text-align: center; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                tr:nth-child(odd) {{ background-color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="company-name">{company_name}</div>
                <div class="report-title">{report_name.replace('_', ' ').title()}</div>
                <div class="date">Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</div>
            </div>
        """
        
        # Process report data
        if isinstance(report_data, dict):
            for section_name, section_data in report_data.items():
                if isinstance(section_data, list) and section_data:
                    html_content += f'<div class="section-title">{section_name.replace("_", " ").title()}</div>'
                    html_content += '<table>'
                    
                    # Add headers
                    headers = list(section_data[0].keys()) if section_data else []
                    html_content += '<tr>'
                    for header in headers:
                        html_content += f'<th>{header.replace("_", " ").title()}</th>'
                    html_content += '</tr>'
                    
                    # Add data rows
                    for row_data in section_data:
                        html_content += '<tr>'
                        for value in row_data.values():
                            html_content += f'<td>{value}</td>'
                        html_content += '</tr>'
                    
                    html_content += '</table>'
        
        elif isinstance(report_data, list) and report_data:
            # Single table format
            html_content += '<table>'
            
            headers = list(report_data[0].keys())
            html_content += '<tr>'
            for header in headers:
                html_content += f'<th>{header.replace("_", " ").title()}</th>'
            html_content += '</tr>'
            
            for row_data in report_data:
                html_content += '<tr>'
                for value in row_data.values():
                    html_content += f'<td>{value}</td>'
                html_content += '</tr>'
            
            html_content += '</table>'
        
        html_content += """
        </body>
        </html>
        """
        
        # Write HTML content to file (which Word can open)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return file_path
    
    def get_download_info(self, file_path):
        """Get download information for a file"""
        if os.path.exists(file_path):
            return {
                'filename': os.path.basename(file_path),
                'size': os.path.getsize(file_path),
                'path': file_path,
                'created': datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            }
        return None