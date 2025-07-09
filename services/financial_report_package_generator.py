"""
Financial Report Package Generator
Creates comprehensive Excel package with all financial reports in separate sheets
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, List, Any

class FinancialReportPackageGenerator:
    """Generates comprehensive Excel package with all financial reports"""
    
    def __init__(self):
        self.report_dir = 'reports_output'
        self.ensure_directories()
    
    def ensure_directories(self):
        """Ensure required directories exist"""
        os.makedirs(self.report_dir, exist_ok=True)
    
    def generate_comprehensive_package(self, company_info: Dict[str, Any] = None) -> str:
        """
        Generate comprehensive Excel package with all financial reports
        Returns path to generated file
        """
        if not company_info:
            company_info = {
                'company_name': 'AccuFin360 Technologies Pvt Ltd',
                'address': '123 Business District, Mumbai, Maharashtra 400001',
                'phone': '+91 22 1234 5678',
                'email': 'info@accufin360.com',
                'reporting_period': 'Financial Year 2024-25',
                'reporting_date': datetime.now().strftime('%B %d, %Y')
            }
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate all financial reports
        self.add_cover_sheet(wb, company_info)
        self.add_balance_sheet(wb, company_info)
        self.add_income_statement(wb, company_info)
        self.add_cash_flow_statement(wb, company_info)
        self.add_statement_of_equity(wb, company_info)
        self.add_trial_balance(wb, company_info)
        self.add_journal_entries(wb, company_info)
        self.add_ledger_summary(wb, company_info)
        self.add_ratio_analysis(wb, company_info)
        self.add_notes_to_financial_statements(wb, company_info)
        self.add_management_discussion(wb, company_info)
        self.add_auditor_report(wb, company_info)
        
        # Save file
        filename = f"Financial_Report_Package_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def add_cover_sheet(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add cover sheet with company information"""
        ws = wb.create_sheet("Cover Sheet")
        
        # Company header
        ws['A1'] = company_info['company_name']
        ws['A1'].font = Font(size=20, bold=True)
        ws['A2'] = "COMPREHENSIVE FINANCIAL REPORT PACKAGE"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = company_info['reporting_period']
        ws['A3'].font = Font(size=12)
        
        # Company details
        ws['A5'] = "Company Information:"
        ws['A5'].font = Font(bold=True)
        ws['A6'] = f"Address: {company_info['address']}"
        ws['A7'] = f"Phone: {company_info['phone']}"
        ws['A8'] = f"Email: {company_info['email']}"
        ws['A9'] = f"Report Date: {company_info['reporting_date']}"
        
        # Table of contents
        ws['A12'] = "TABLE OF CONTENTS"
        ws['A12'].font = Font(size=14, bold=True)
        
        reports = [
            "1. Balance Sheet",
            "2. Income Statement", 
            "3. Cash Flow Statement",
            "4. Statement of Equity",
            "5. Trial Balance",
            "6. Journal Entries",
            "7. Ledger Summary",
            "8. Ratio Analysis",
            "9. Notes to Financial Statements",
            "10. Management Discussion",
            "11. Auditor Report"
        ]
        
        for i, report in enumerate(reports, 13):
            ws[f'A{i}'] = report
        
        # Compliance note
        ws['A26'] = "COMPLIANCE STATEMENT"
        ws['A26'].font = Font(bold=True)
        ws['A27'] = "These financial statements have been prepared in accordance with:"
        ws['A28'] = "• International Financial Reporting Standards (IFRS)"
        ws['A29'] = "• Generally Accepted Accounting Principles (GAAP)"
        ws['A30'] = "• Indian Accounting Standards (Ind AS)"
        ws['A31'] = "• Companies Act, 2013"
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def add_balance_sheet(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Balance Sheet"""
        ws = wb.create_sheet("Balance Sheet")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "BALANCE SHEET"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"As at {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Balance sheet data
        balance_sheet_data = [
            ["", "Amount (₹)", ""],
            ["ASSETS", "", ""],
            ["", "", ""],
            ["Current Assets:", "", ""],
            ["Cash and Cash Equivalents", "23,750,000", ""],
            ["Trade Receivables", "13,750,000", ""],
            ["Inventory", "7,425,000", ""],
            ["Prepaid Expenses", "1,500,000", ""],
            ["Other Current Assets", "2,575,000", ""],
            ["Total Current Assets", "49,000,000", ""],
            ["", "", ""],
            ["Non-Current Assets:", "", ""],
            ["Property, Plant & Equipment", "35,000,000", ""],
            ["Intangible Assets", "8,500,000", ""],
            ["Investments", "12,000,000", ""],
            ["Deferred Tax Assets", "1,750,000", ""],
            ["Other Non-Current Assets", "3,250,000", ""],
            ["Total Non-Current Assets", "60,500,000", ""],
            ["", "", ""],
            ["TOTAL ASSETS", "109,500,000", ""],
            ["", "", ""],
            ["LIABILITIES AND EQUITY", "", ""],
            ["", "", ""],
            ["Current Liabilities:", "", ""],
            ["Trade Payables", "8,250,000", ""],
            ["Other Current Liabilities", "3,100,000", ""],
            ["Short-term Borrowings", "5,000,000", ""],
            ["Current Portion of Long-term Debt", "2,500,000", ""],
            ["Provisions", "1,150,000", ""],
            ["Total Current Liabilities", "20,000,000", ""],
            ["", "", ""],
            ["Non-Current Liabilities:", "", ""],
            ["Long-term Borrowings", "15,000,000", ""],
            ["Deferred Tax Liabilities", "2,250,000", ""],
            ["Employee Benefits", "1,500,000", ""],
            ["Other Non-Current Liabilities", "750,000", ""],
            ["Total Non-Current Liabilities", "19,500,000", ""],
            ["", "", ""],
            ["Total Liabilities", "39,500,000", ""],
            ["", "", ""],
            ["EQUITY:", "", ""],
            ["Share Capital", "25,000,000", ""],
            ["Reserves and Surplus", "35,000,000", ""],
            ["Retained Earnings", "10,000,000", ""],
            ["Total Equity", "70,000,000", ""],
            ["", "", ""],
            ["TOTAL LIABILITIES AND EQUITY", "109,500,000", ""]
        ]
        
        # Write data
        for i, row in enumerate(balance_sheet_data, 5):
            ws[f'A{i}'] = row[0]
            ws[f'B{i}'] = row[1]
            
            # Format headers and totals
            if any(keyword in str(row[0]).upper() for keyword in ['ASSETS', 'LIABILITIES', 'EQUITY:']):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
            elif 'Total' in str(row[0]):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def add_income_statement(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Income Statement"""
        ws = wb.create_sheet("Income Statement")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "INCOME STATEMENT"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Income statement data
        income_data = [
            ["", "Amount (₹)", ""],
            ["REVENUE:", "", ""],
            ["Sales Revenue", "125,000,000", ""],
            ["Service Revenue", "45,000,000", ""],
            ["Other Income", "3,750,000", ""],
            ["Total Revenue", "173,750,000", ""],
            ["", "", ""],
            ["COST OF GOODS SOLD:", "", ""],
            ["Raw Materials", "45,000,000", ""],
            ["Direct Labor", "20,000,000", ""],
            ["Manufacturing Overheads", "22,500,000", ""],
            ["Total Cost of Goods Sold", "87,500,000", ""],
            ["", "", ""],
            ["GROSS PROFIT", "86,250,000", ""],
            ["", "", ""],
            ["OPERATING EXPENSES:", "", ""],
            ["Salaries and Wages", "25,000,000", ""],
            ["Rent and Utilities", "12,500,000", ""],
            ["Marketing and Advertising", "8,750,000", ""],
            ["Professional Fees", "3,250,000", ""],
            ["Depreciation", "5,000,000", ""],
            ["Other Operating Expenses", "4,500,000", ""],
            ["Total Operating Expenses", "59,000,000", ""],
            ["", "", ""],
            ["OPERATING PROFIT", "27,250,000", ""],
            ["", "", ""],
            ["OTHER INCOME/(EXPENSES):", "", ""],
            ["Interest Income", "2,500,000", ""],
            ["Interest Expense", "(3,750,000)", ""],
            ["Other Income", "1,250,000", ""],
            ["Total Other Income/(Expenses)", "0", ""],
            ["", "", ""],
            ["PROFIT BEFORE TAX", "27,250,000", ""],
            ["", "", ""],
            ["TAX EXPENSES:", "", ""],
            ["Current Tax", "8,175,000", ""],
            ["Deferred Tax", "1,075,000", ""],
            ["Total Tax Expenses", "9,250,000", ""],
            ["", "", ""],
            ["NET PROFIT", "18,000,000", ""],
            ["", "", ""],
            ["EARNINGS PER SHARE:", "", ""],
            ["Basic EPS (₹)", "7.20", ""],
            ["Diluted EPS (₹)", "7.20", ""]
        ]
        
        # Write data
        for i, row in enumerate(income_data, 5):
            ws[f'A{i}'] = row[0]
            ws[f'B{i}'] = row[1]
            
            # Format headers and totals
            if any(keyword in str(row[0]).upper() for keyword in ['REVENUE:', 'COST OF', 'OPERATING', 'OTHER INCOME', 'TAX', 'NET PROFIT', 'EARNINGS']):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
            elif 'Total' in str(row[0]) or 'GROSS PROFIT' in str(row[0]) or 'PROFIT BEFORE' in str(row[0]):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def add_cash_flow_statement(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Cash Flow Statement"""
        ws = wb.create_sheet("Cash Flow Statement")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "CASH FLOW STATEMENT"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Cash flow data
        cash_flow_data = [
            ["", "Amount (₹)", ""],
            ["CASH FLOWS FROM OPERATING ACTIVITIES:", "", ""],
            ["Net Profit", "18,000,000", ""],
            ["Adjustments for:", "", ""],
            ["Depreciation", "5,000,000", ""],
            ["Interest Expense", "3,750,000", ""],
            ["Interest Income", "(2,500,000)", ""],
            ["Provision for Bad Debts", "750,000", ""],
            ["", "", ""],
            ["Operating Profit before Working Capital Changes", "25,000,000", ""],
            ["", "", ""],
            ["Changes in Working Capital:", "", ""],
            ["Trade Receivables", "(2,500,000)", ""],
            ["Inventory", "(1,250,000)", ""],
            ["Trade Payables", "1,750,000", ""],
            ["Other Current Assets/Liabilities", "500,000", ""],
            ["", "", ""],
            ["Cash Generated from Operations", "23,500,000", ""],
            ["Interest Paid", "(3,750,000)", ""],
            ["Income Tax Paid", "(8,175,000)", ""],
            ["Net Cash from Operating Activities", "11,575,000", ""],
            ["", "", ""],
            ["CASH FLOWS FROM INVESTING ACTIVITIES:", "", ""],
            ["Purchase of Property, Plant & Equipment", "(15,000,000)", ""],
            ["Purchase of Investments", "(5,000,000)", ""],
            ["Sale of Investments", "3,000,000", ""],
            ["Interest Received", "2,500,000", ""],
            ["Net Cash used in Investing Activities", "(14,500,000)", ""],
            ["", "", ""],
            ["CASH FLOWS FROM FINANCING ACTIVITIES:", "", ""],
            ["Proceeds from Share Issue", "5,000,000", ""],
            ["Proceeds from Borrowings", "10,000,000", ""],
            ["Repayment of Borrowings", "(7,500,000)", ""],
            ["Dividend Paid", "(2,500,000)", ""],
            ["Net Cash from Financing Activities", "5,000,000", ""],
            ["", "", ""],
            ["NET INCREASE IN CASH AND CASH EQUIVALENTS", "2,075,000", ""],
            ["Cash and Cash Equivalents at Beginning", "21,675,000", ""],
            ["Cash and Cash Equivalents at End", "23,750,000", ""]
        ]
        
        # Write data
        for i, row in enumerate(cash_flow_data, 5):
            ws[f'A{i}'] = row[0]
            ws[f'B{i}'] = row[1]
            
            # Format headers and totals
            if any(keyword in str(row[0]).upper() for keyword in ['CASH FLOWS FROM', 'NET CASH', 'NET INCREASE', 'CASH AND CASH EQUIVALENTS']):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
    
    def add_statement_of_equity(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Statement of Changes in Equity"""
        ws = wb.create_sheet("Statement of Equity")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "STATEMENT OF CHANGES IN EQUITY"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Equity statement data
        equity_data = [
            ["", "Share Capital", "Reserves", "Retained Earnings", "Total Equity"],
            ["", "(₹)", "(₹)", "(₹)", "(₹)"],
            ["", "", "", "", ""],
            ["Balance at Beginning of Year", "20,000,000", "30,000,000", "12,500,000", "62,500,000"],
            ["", "", "", "", ""],
            ["Changes during the year:", "", "", "", ""],
            ["Issue of Share Capital", "5,000,000", "0", "0", "5,000,000"],
            ["Transfer to Reserves", "0", "5,000,000", "(5,000,000)", "0"],
            ["Net Profit for the Year", "0", "0", "18,000,000", "18,000,000"],
            ["Dividend Declared", "0", "0", "(2,500,000)", "(2,500,000)"],
            ["Other Comprehensive Income", "0", "0", "500,000", "500,000"],
            ["", "", "", "", ""],
            ["Total Changes", "5,000,000", "5,000,000", "11,000,000", "21,000,000"],
            ["", "", "", "", ""],
            ["Balance at End of Year", "25,000,000", "35,000,000", "23,500,000", "83,500,000"]
        ]
        
        # Write data
        for i, row in enumerate(equity_data, 5):
            for j, cell_value in enumerate(row):
                ws.cell(row=i, column=j+1, value=cell_value)
                
                # Format headers and totals
                if i == 5 or i == 6:  # Header rows
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
                elif 'Balance' in str(row[0]) or 'Total' in str(row[0]):
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
    
    def add_trial_balance(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Trial Balance"""
        ws = wb.create_sheet("Trial Balance")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "TRIAL BALANCE"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"As at {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Trial balance data
        trial_balance_data = [
            ["Account Name", "Debit (₹)", "Credit (₹)"],
            ["", "", ""],
            ["Cash and Bank", "23,750,000", ""],
            ["Trade Receivables", "13,750,000", ""],
            ["Inventory", "7,425,000", ""],
            ["Prepaid Expenses", "1,500,000", ""],
            ["Property, Plant & Equipment", "35,000,000", ""],
            ["Intangible Assets", "8,500,000", ""],
            ["Investments", "12,000,000", ""],
            ["Cost of Goods Sold", "87,500,000", ""],
            ["Salaries and Wages", "25,000,000", ""],
            ["Rent and Utilities", "12,500,000", ""],
            ["Marketing Expenses", "8,750,000", ""],
            ["Professional Fees", "3,250,000", ""],
            ["Depreciation", "5,000,000", ""],
            ["Interest Expense", "3,750,000", ""],
            ["Income Tax Expense", "9,250,000", ""],
            ["Trade Payables", "", "8,250,000"],
            ["Other Current Liabilities", "", "3,100,000"],
            ["Short-term Borrowings", "", "5,000,000"],
            ["Long-term Borrowings", "", "15,000,000"],
            ["Share Capital", "", "25,000,000"],
            ["Reserves and Surplus", "", "35,000,000"],
            ["Retained Earnings", "", "10,000,000"],
            ["Sales Revenue", "", "125,000,000"],
            ["Service Revenue", "", "45,000,000"],
            ["Other Income", "", "3,750,000"],
            ["Interest Income", "", "2,500,000"],
            ["", "", ""],
            ["TOTAL", "257,925,000", "257,925,000"]
        ]
        
        # Write data
        for i, row in enumerate(trial_balance_data, 5):
            for j, cell_value in enumerate(row):
                ws.cell(row=i, column=j+1, value=cell_value)
                
                # Format headers and totals
                if i == 5 or row[0] == 'TOTAL':  # Header row and total
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def add_journal_entries(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Journal Entries Summary"""
        ws = wb.create_sheet("Journal Entries")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "JOURNAL ENTRIES SUMMARY"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Journal entries data
        journal_data = [
            ["Entry No.", "Date", "Account", "Description", "Debit (₹)", "Credit (₹)"],
            ["", "", "", "", "", ""],
            ["JE001", "2024-01-01", "Cash", "Opening Balance", "21,675,000", ""],
            ["", "", "Capital", "Opening Balance", "", "20,000,000"],
            ["", "", "Retained Earnings", "Opening Balance", "", "1,675,000"],
            ["", "", "", "", "", ""],
            ["JE002", "2024-01-15", "Trade Receivables", "Sales on Credit", "25,000,000", ""],
            ["", "", "Sales Revenue", "Sales on Credit", "", "25,000,000"],
            ["", "", "", "", "", ""],
            ["JE003", "2024-01-20", "Inventory", "Purchase of Goods", "7,425,000", ""],
            ["", "", "Trade Payables", "Purchase of Goods", "", "7,425,000"],
            ["", "", "", "", "", ""],
            ["JE004", "2024-02-01", "Salaries Expense", "Monthly Salaries", "2,500,000", ""],
            ["", "", "Cash", "Monthly Salaries", "", "2,500,000"],
            ["", "", "", "", "", ""],
            ["JE005", "2024-02-15", "Rent Expense", "Office Rent", "125,000", ""],
            ["", "", "Cash", "Office Rent", "", "125,000"],
            ["", "", "", "", "", ""],
            ["JE006", "2024-03-01", "Interest Expense", "Loan Interest", "312,500", ""],
            ["", "", "Cash", "Loan Interest", "", "312,500"],
            ["", "", "", "", "", ""],
            ["JE007", "2024-03-31", "Depreciation Expense", "Monthly Depreciation", "416,667", ""],
            ["", "", "Accumulated Depreciation", "Monthly Depreciation", "", "416,667"]
        ]
        
        # Write data
        for i, row in enumerate(journal_data, 5):
            for j, cell_value in enumerate(row):
                ws.cell(row=i, column=j+1, value=cell_value)
                
                # Format header
                if i == 5:  # Header row
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def add_ledger_summary(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Ledger Summary"""
        ws = wb.create_sheet("Ledger Summary")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "LEDGER SUMMARY"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Ledger summary data
        ledger_data = [
            ["Account Name", "Account Type", "Opening Balance", "Debit", "Credit", "Closing Balance"],
            ["", "", "(₹)", "(₹)", "(₹)", "(₹)"],
            ["", "", "", "", "", ""],
            ["Cash and Bank", "Asset", "21,675,000", "5,000,000", "2,925,000", "23,750,000"],
            ["Trade Receivables", "Asset", "11,250,000", "27,500,000", "25,000,000", "13,750,000"],
            ["Inventory", "Asset", "6,175,000", "8,675,000", "7,425,000", "7,425,000"],
            ["Property, Plant & Equipment", "Asset", "20,000,000", "15,000,000", "0", "35,000,000"],
            ["Trade Payables", "Liability", "6,500,000", "6,675,000", "8,425,000", "8,250,000"],
            ["Short-term Borrowings", "Liability", "3,000,000", "0", "2,000,000", "5,000,000"],
            ["Long-term Borrowings", "Liability", "5,000,000", "0", "10,000,000", "15,000,000"],
            ["Share Capital", "Equity", "20,000,000", "0", "5,000,000", "25,000,000"],
            ["Sales Revenue", "Revenue", "0", "0", "125,000,000", "125,000,000"],
            ["Service Revenue", "Revenue", "0", "0", "45,000,000", "45,000,000"],
            ["Cost of Goods Sold", "Expense", "0", "87,500,000", "0", "87,500,000"],
            ["Salaries and Wages", "Expense", "0", "25,000,000", "0", "25,000,000"],
            ["Rent and Utilities", "Expense", "0", "12,500,000", "0", "12,500,000"],
            ["Marketing Expenses", "Expense", "0", "8,750,000", "0", "8,750,000"],
            ["Interest Expense", "Expense", "0", "3,750,000", "0", "3,750,000"]
        ]
        
        # Write data
        for i, row in enumerate(ledger_data, 5):
            for j, cell_value in enumerate(row):
                ws.cell(row=i, column=j+1, value=cell_value)
                
                # Format headers
                if i == 5 or i == 6:  # Header rows
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
    
    def add_ratio_analysis(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Financial Ratio Analysis"""
        ws = wb.create_sheet("Ratio Analysis")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "FINANCIAL RATIO ANALYSIS"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Ratio analysis data
        ratio_data = [
            ["Ratio Category", "Ratio Name", "Formula", "Value", "Benchmark"],
            ["", "", "", "", ""],
            ["Liquidity Ratios", "", "", "", ""],
            ["", "Current Ratio", "Current Assets / Current Liabilities", "2.45", "1.5-2.0"],
            ["", "Quick Ratio", "Quick Assets / Current Liabilities", "1.95", "1.0-1.5"],
            ["", "Cash Ratio", "Cash / Current Liabilities", "1.19", "0.2-0.5"],
            ["", "", "", "", ""],
            ["Activity Ratios", "", "", "", ""],
            ["", "Asset Turnover", "Sales / Total Assets", "1.14", "1.0-2.0"],
            ["", "Inventory Turnover", "COGS / Inventory", "11.78", "6.0-12.0"],
            ["", "Receivables Turnover", "Sales / Receivables", "9.09", "8.0-15.0"],
            ["", "", "", "", ""],
            ["Profitability Ratios", "", "", "", ""],
            ["", "Gross Profit Margin", "Gross Profit / Sales", "49.64%", "20%-50%"],
            ["", "Operating Profit Margin", "Operating Profit / Sales", "15.69%", "5%-20%"],
            ["", "Net Profit Margin", "Net Profit / Sales", "10.36%", "5%-15%"],
            ["", "Return on Assets", "Net Profit / Total Assets", "16.44%", "5%-20%"],
            ["", "Return on Equity", "Net Profit / Equity", "25.71%", "10%-25%"],
            ["", "", "", "", ""],
            ["Leverage Ratios", "", "", "", ""],
            ["", "Debt-to-Equity", "Total Debt / Equity", "0.56", "0.3-0.7"],
            ["", "Debt-to-Assets", "Total Debt / Total Assets", "0.36", "0.3-0.6"],
            ["", "Interest Coverage", "EBIT / Interest Expense", "7.27", "2.5-5.0"],
            ["", "", "", "", ""],
            ["Market Ratios", "", "", "", ""],
            ["", "Earnings per Share", "Net Profit / Shares Outstanding", "₹7.20", "Industry Avg"],
            ["", "Book Value per Share", "Equity / Shares Outstanding", "₹28.00", "Industry Avg"]
        ]
        
        # Write data
        for i, row in enumerate(ratio_data, 5):
            for j, cell_value in enumerate(row):
                ws.cell(row=i, column=j+1, value=cell_value)
                
                # Format headers and categories
                if i == 5 or any(category in str(row[0]) for category in ['Liquidity', 'Activity', 'Profitability', 'Leverage', 'Market']):
                    ws.cell(row=i, column=j+1).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def add_notes_to_financial_statements(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Notes to Financial Statements"""
        ws = wb.create_sheet("Notes")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "NOTES TO FINANCIAL STATEMENTS"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Notes data
        notes_data = [
            ["Note 1: Basis of Preparation"],
            ["These financial statements have been prepared in accordance with"],
            ["International Financial Reporting Standards (IFRS) and Indian Accounting Standards."],
            [""],
            ["Note 2: Significant Accounting Policies"],
            ["a) Revenue Recognition: Revenue is recognized when control of goods or services"],
            ["   is transferred to customers at the amount expected to be received."],
            ["b) Inventory Valuation: Inventories are valued at lower of cost or net realizable value."],
            ["c) Depreciation: Depreciation is provided on straight-line method."],
            [""],
            ["Note 3: Property, Plant & Equipment"],
            ["Gross Block: ₹50,000,000"],
            ["Accumulated Depreciation: ₹15,000,000"],
            ["Net Block: ₹35,000,000"],
            [""],
            ["Note 4: Trade Receivables"],
            ["Less than 6 months: ₹12,000,000"],
            ["6 months to 1 year: ₹1,500,000"],
            ["More than 1 year: ₹250,000"],
            ["Total: ₹13,750,000"],
            [""],
            ["Note 5: Borrowings"],
            ["Term Loans from Banks: ₹15,000,000"],
            ["Working Capital Loans: ₹5,000,000"],
            ["Total: ₹20,000,000"],
            [""],
            ["Note 6: Contingent Liabilities"],
            ["Bank Guarantees: ₹2,500,000"],
            ["Legal Cases: ₹1,000,000"],
            [""],
            ["Note 7: Subsequent Events"],
            ["No material events have occurred after the balance sheet date"],
            ["that would require adjustment to these financial statements."]
        ]
        
        # Write data
        for i, row in enumerate(notes_data, 5):
            ws[f'A{i}'] = row[0]
            
            # Format note headers
            if row[0].startswith('Note'):
                ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 70
    
    def add_management_discussion(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Management Discussion and Analysis"""
        ws = wb.create_sheet("Management Discussion")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "MANAGEMENT DISCUSSION AND ANALYSIS"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Management discussion data
        md_data = [
            ["EXECUTIVE SUMMARY"],
            ["The Company has delivered strong financial performance during the year"],
            ["with revenue growth of 15% and improved profitability margins."],
            [""],
            ["FINANCIAL PERFORMANCE"],
            ["• Total Revenue: ₹173.75 crores (15% growth)"],
            ["• Gross Profit Margin: 49.64% (improved from 47.5%)"],
            ["• Net Profit: ₹18.00 crores (22% growth)"],
            ["• Return on Equity: 25.71% (industry-leading performance)"],
            [""],
            ["OPERATIONAL HIGHLIGHTS"],
            ["• Expanded market presence in 3 new geographic regions"],
            ["• Launched 2 new product lines contributing 12% to revenue"],
            ["• Improved operational efficiency through automation initiatives"],
            ["• Enhanced customer satisfaction scores to 4.2/5.0"],
            [""],
            ["MARKET POSITION"],
            ["The Company maintains a strong market position with:"],
            ["• Market share of 8.5% in core business segment"],
            ["• Strong brand recognition and customer loyalty"],
            ["• Diversified customer base across industries"],
            [""],
            ["RISK FACTORS"],
            ["• Economic uncertainty and market volatility"],
            ["• Regulatory changes in key markets"],
            ["• Competition from new market entrants"],
            ["• Supply chain disruptions"],
            [""],
            ["OUTLOOK"],
            ["Management expects continued growth driven by:"],
            ["• New product launches planned for next year"],
            ["• Expansion into emerging markets"],
            ["• Strategic partnerships and acquisitions"],
            ["• Investment in technology and innovation"],
            [""],
            ["CONCLUSION"],
            ["The Company is well-positioned for sustainable growth with"],
            ["strong fundamentals and strategic initiatives in place."]
        ]
        
        # Write data
        for i, row in enumerate(md_data, 5):
            ws[f'A{i}'] = row[0]
            
            # Format section headers
            if row[0] in ['EXECUTIVE SUMMARY', 'FINANCIAL PERFORMANCE', 'OPERATIONAL HIGHLIGHTS', 
                         'MARKET POSITION', 'RISK FACTORS', 'OUTLOOK', 'CONCLUSION']:
                ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 70
    
    def add_auditor_report(self, wb: Workbook, company_info: Dict[str, Any]):
        """Add Independent Auditor's Report"""
        ws = wb.create_sheet("Auditor Report")
        
        # Header
        ws['A1'] = f"{company_info['company_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "INDEPENDENT AUDITOR'S REPORT"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A3'] = f"For the year ended {company_info['reporting_date']}"
        ws['A3'].font = Font(size=12)
        
        # Auditor report data
        audit_data = [
            ["TO THE MEMBERS OF " + company_info['company_name'].upper()],
            [""],
            ["REPORT ON THE AUDIT OF FINANCIAL STATEMENTS"],
            [""],
            ["Opinion"],
            ["We have audited the financial statements of " + company_info['company_name'] + ","],
            ["which comprise the balance sheet as at " + company_info['reporting_date'] + ","],
            ["and the statement of profit and loss, statement of changes in equity and"],
            ["statement of cash flows for the year then ended, and notes to the financial"],
            ["statements, including a summary of significant accounting policies."],
            [""],
            ["In our opinion and to the best of our information and according to the"],
            ["explanations given to us, the aforesaid financial statements give the"],
            ["information required by the Act in the manner so required and give a"],
            ["true and fair view in conformity with the accounting principles generally"],
            ["accepted in India, of the state of affairs of the Company as at"],
            [company_info['reporting_date'] + " and profit and cash flows for the year ended on that date."],
            [""],
            ["Basis for Opinion"],
            ["We conducted our audit in accordance with the Standards on Auditing (SAs)"],
            ["specified under section 143(10) of the Companies Act, 2013. Our responsibilities"],
            ["under those Standards are further described in the Auditor's Responsibilities"],
            ["for the Audit of the Financial Statements section of our report."],
            [""],
            ["Key Audit Matters"],
            ["Key audit matters are those matters that, in our professional judgment,"],
            ["were of most significance in our audit of the financial statements"],
            ["of the current period. These matters were addressed in the context of"],
            ["our audit of the financial statements as a whole."],
            [""],
            ["Management's Responsibility for the Financial Statements"],
            ["The Company's management is responsible for the preparation and fair"],
            ["presentation of these financial statements in accordance with the accounting"],
            ["principles generally accepted in India."],
            [""],
            ["Auditor's Responsibility"],
            ["Our responsibility is to express an opinion on these financial statements"],
            ["based on our audit. We conducted our audit in accordance with the Standards"],
            ["on Auditing specified under Section 143(10) of the Companies Act, 2013."],
            [""],
            [""],
            ["For ABC & Associates"],
            ["Chartered Accountants"],
            ["FRN: 012345N"],
            [""],
            [""],
            ["(Partner Name)"],
            ["Partner"],
            ["Membership No: 123456"],
            [""],
            ["Place: Mumbai"],
            ["Date: " + company_info['reporting_date']]
        ]
        
        # Write data
        for i, row in enumerate(audit_data, 5):
            ws[f'A{i}'] = row[0]
            
            # Format headers
            if row[0] in ['Opinion', 'Basis for Opinion', 'Key Audit Matters', 
                         "Management's Responsibility for the Financial Statements", 
                         "Auditor's Responsibility", 'REPORT ON THE AUDIT OF FINANCIAL STATEMENTS']:
                ws[f'A{i}'].font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 80
    
    def format_currency(self, value: float) -> str:
        """Format currency with Indian Rupee symbol"""
        return f"₹{value:,.0f}"

# Test the generator
if __name__ == "__main__":
    generator = FinancialReportPackageGenerator()
    filepath = generator.generate_comprehensive_package()
    print(f"Generated comprehensive financial report package: {filepath}")