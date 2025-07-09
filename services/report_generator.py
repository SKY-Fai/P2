import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from app import db
from models import JournalEntry, ChartOfAccount, FinancialReport, Company
from services.accounting_engine import AccountingEngine

class ReportGenerator:
    """Generate financial reports in multiple formats"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.accounting_engine = AccountingEngine()
        self.company_id = 1  # Default company ID
    
    def generate_all_reports(self, file_id: int) -> Dict[str, Any]:
        """Generate all standard financial reports"""
        try:
            reports = {}
            
            # Generate each report type
            reports['journal_entries'] = self.generate_journal_entries_report(file_id)
            reports['trial_balance'] = self.generate_trial_balance_report(file_id)
            reports['balance_sheet'] = self.generate_balance_sheet_report(file_id)
            reports['income_statement'] = self.generate_income_statement_report(file_id)
            reports['cash_flow'] = self.generate_cash_flow_report(file_id)
            reports['ledger_summary'] = self.generate_ledger_summary_report(file_id)
            
            return reports
            
        except Exception as e:
            self.logger.error(f"Error generating reports: {str(e)}")
            raise
    
    def generate_journal_entries_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Journal Entries Report"""
        try:
            # Get journal entries
            entries = JournalEntry.query.filter_by(company_id=self.company_id).order_by(JournalEntry.entry_date.desc()).all()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal Entries"
            
            # Headers
            headers = ['Date', 'Description', 'Account Code', 'Account Name', 'Debit', 'Credit', 'Reference']
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for entry in entries:
                row_data = [
                    entry.entry_date.strftime('%Y-%m-%d'),
                    entry.description,
                    entry.account.account_code,
                    entry.account.account_name,
                    entry.debit_amount if entry.debit_amount > 0 else '',
                    entry.credit_amount if entry.credit_amount > 0 else '',
                    entry.reference_number or ''
                ]
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"journal_entries_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='journal_entries',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'journal_entries',
                'filename': filename,
                'filepath': filepath,
                'total_entries': len(entries),
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating journal entries report: {str(e)}")
            raise
    
    def generate_trial_balance_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Trial Balance Report"""
        try:
            # Get trial balance data
            trial_balance = self.accounting_engine.get_trial_balance()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            
            # Title
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = "TRIAL BALANCE"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Date range
            ws.merge_cells('A2:D2')
            date_cell = ws['A2']
            date_cell.value = f"From {trial_balance['date_from']} to {trial_balance['date_to']}"
            date_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
            ws.append([''] * 4)  # Empty row
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Add account data
            for account in trial_balance['accounts']:
                debit_amount = account['balance'] if account['balance'] > 0 and account['debit_total'] > account['credit_total'] else ''
                credit_amount = account['balance'] if account['balance'] > 0 and account['credit_total'] > account['debit_total'] else ''
                
                row_data = [
                    account['account_code'],
                    account['account_name'],
                    debit_amount,
                    credit_amount
                ]
                ws.append(row_data)
            
            # Add totals
            ws.append(['', 'TOTAL', trial_balance['total_debits'], trial_balance['total_credits']])
            
            # Style totals row
            total_row = ws.max_row
            for col_num in range(1, 5):
                cell = ws.cell(row=total_row, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"trial_balance_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='trial_balance',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'trial_balance',
                'filename': filename,
                'filepath': filepath,
                'total_debits': trial_balance['total_debits'],
                'total_credits': trial_balance['total_credits'],
                'is_balanced': trial_balance['is_balanced'],
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating trial balance report: {str(e)}")
            raise
    
    def generate_balance_sheet_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Balance Sheet Report"""
        try:
            # Get balance sheet data
            balance_sheet_data = self._get_balance_sheet_data()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance Sheet"
            
            # Title
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "BALANCE SHEET"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Date
            ws.merge_cells('A2:C2')
            date_cell = ws['A2']
            date_cell.value = f"As of {datetime.now().strftime('%B %d, %Y')}"
            date_cell.alignment = Alignment(horizontal='center')
            
            ws.append([''] * 3)  # Empty row
            
            # ASSETS
            ws.append(['ASSETS', '', ''])
            assets_row = ws.max_row
            ws.cell(row=assets_row, column=1).font = Font(bold=True)
            
            current_assets_total = 0
            for asset in balance_sheet_data['current_assets']:
                ws.append([f"  {asset['name']}", '', asset['amount']])
                current_assets_total += asset['amount']
            
            ws.append(['  Total Current Assets', '', current_assets_total])
            current_assets_row = ws.max_row
            ws.cell(row=current_assets_row, column=1).font = Font(bold=True)
            
            fixed_assets_total = 0
            for asset in balance_sheet_data['fixed_assets']:
                ws.append([f"  {asset['name']}", '', asset['amount']])
                fixed_assets_total += asset['amount']
            
            ws.append(['  Total Fixed Assets', '', fixed_assets_total])
            fixed_assets_row = ws.max_row
            ws.cell(row=fixed_assets_row, column=1).font = Font(bold=True)
            
            total_assets = current_assets_total + fixed_assets_total
            ws.append(['TOTAL ASSETS', '', total_assets])
            total_assets_row = ws.max_row
            ws.cell(row=total_assets_row, column=1).font = Font(bold=True)
            ws.cell(row=total_assets_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # LIABILITIES
            ws.append(['LIABILITIES', '', ''])
            liabilities_row = ws.max_row
            ws.cell(row=liabilities_row, column=1).font = Font(bold=True)
            
            current_liabilities_total = 0
            for liability in balance_sheet_data['current_liabilities']:
                ws.append([f"  {liability['name']}", '', liability['amount']])
                current_liabilities_total += liability['amount']
            
            ws.append(['  Total Current Liabilities', '', current_liabilities_total])
            current_liabilities_row = ws.max_row
            ws.cell(row=current_liabilities_row, column=1).font = Font(bold=True)
            
            long_term_liabilities_total = 0
            for liability in balance_sheet_data['long_term_liabilities']:
                ws.append([f"  {liability['name']}", '', liability['amount']])
                long_term_liabilities_total += liability['amount']
            
            ws.append(['  Total Long-term Liabilities', '', long_term_liabilities_total])
            long_term_liabilities_row = ws.max_row
            ws.cell(row=long_term_liabilities_row, column=1).font = Font(bold=True)
            
            total_liabilities = current_liabilities_total + long_term_liabilities_total
            ws.append(['TOTAL LIABILITIES', '', total_liabilities])
            total_liabilities_row = ws.max_row
            ws.cell(row=total_liabilities_row, column=1).font = Font(bold=True)
            ws.cell(row=total_liabilities_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # EQUITY
            ws.append(['EQUITY', '', ''])
            equity_row = ws.max_row
            ws.cell(row=equity_row, column=1).font = Font(bold=True)
            
            total_equity = 0
            for equity in balance_sheet_data['equity']:
                ws.append([f"  {equity['name']}", '', equity['amount']])
                total_equity += equity['amount']
            
            ws.append(['TOTAL EQUITY', '', total_equity])
            total_equity_row = ws.max_row
            ws.cell(row=total_equity_row, column=1).font = Font(bold=True)
            ws.cell(row=total_equity_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # TOTAL LIABILITIES AND EQUITY
            total_liabilities_equity = total_liabilities + total_equity
            ws.append(['TOTAL LIABILITIES AND EQUITY', '', total_liabilities_equity])
            total_row = ws.max_row
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=3).font = Font(bold=True)
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"balance_sheet_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='balance_sheet',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'balance_sheet',
                'filename': filename,
                'filepath': filepath,
                'total_assets': total_assets,
                'total_liabilities': total_liabilities,
                'total_equity': total_equity,
                'is_balanced': abs(total_assets - total_liabilities_equity) < 0.01,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating balance sheet report: {str(e)}")
            raise
    
    def generate_income_statement_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Income Statement (P&L) Report"""
        try:
            # Get income statement data
            income_statement_data = self._get_income_statement_data()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Income Statement"
            
            # Title
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "INCOME STATEMENT"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Date range
            ws.merge_cells('A2:C2')
            date_cell = ws['A2']
            date_cell.value = f"For the period ending {datetime.now().strftime('%B %d, %Y')}"
            date_cell.alignment = Alignment(horizontal='center')
            
            ws.append([''] * 3)  # Empty row
            
            # REVENUE
            ws.append(['REVENUE', '', ''])
            revenue_row = ws.max_row
            ws.cell(row=revenue_row, column=1).font = Font(bold=True)
            
            total_revenue = 0
            for revenue in income_statement_data['revenue']:
                ws.append([f"  {revenue['name']}", '', revenue['amount']])
                total_revenue += revenue['amount']
            
            ws.append(['TOTAL REVENUE', '', total_revenue])
            total_revenue_row = ws.max_row
            ws.cell(row=total_revenue_row, column=1).font = Font(bold=True)
            ws.cell(row=total_revenue_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # COST OF GOODS SOLD
            ws.append(['COST OF GOODS SOLD', '', ''])
            cogs_row = ws.max_row
            ws.cell(row=cogs_row, column=1).font = Font(bold=True)
            
            total_cogs = 0
            for cogs in income_statement_data['cogs']:
                ws.append([f"  {cogs['name']}", '', cogs['amount']])
                total_cogs += cogs['amount']
            
            ws.append(['TOTAL COST OF GOODS SOLD', '', total_cogs])
            total_cogs_row = ws.max_row
            ws.cell(row=total_cogs_row, column=1).font = Font(bold=True)
            ws.cell(row=total_cogs_row, column=3).font = Font(bold=True)
            
            # GROSS PROFIT
            gross_profit = total_revenue - total_cogs
            ws.append(['GROSS PROFIT', '', gross_profit])
            gross_profit_row = ws.max_row
            ws.cell(row=gross_profit_row, column=1).font = Font(bold=True)
            ws.cell(row=gross_profit_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # OPERATING EXPENSES
            ws.append(['OPERATING EXPENSES', '', ''])
            expenses_row = ws.max_row
            ws.cell(row=expenses_row, column=1).font = Font(bold=True)
            
            total_expenses = 0
            for expense in income_statement_data['expenses']:
                ws.append([f"  {expense['name']}", '', expense['amount']])
                total_expenses += expense['amount']
            
            ws.append(['TOTAL OPERATING EXPENSES', '', total_expenses])
            total_expenses_row = ws.max_row
            ws.cell(row=total_expenses_row, column=1).font = Font(bold=True)
            ws.cell(row=total_expenses_row, column=3).font = Font(bold=True)
            
            # NET INCOME
            net_income = gross_profit - total_expenses
            ws.append(['NET INCOME', '', net_income])
            net_income_row = ws.max_row
            ws.cell(row=net_income_row, column=1).font = Font(bold=True)
            ws.cell(row=net_income_row, column=3).font = Font(bold=True)
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"income_statement_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='income_statement',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'income_statement',
                'filename': filename,
                'filepath': filepath,
                'total_revenue': total_revenue,
                'total_expenses': total_expenses,
                'gross_profit': gross_profit,
                'net_income': net_income,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating income statement report: {str(e)}")
            raise
    
    def generate_cash_flow_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Cash Flow Statement"""
        try:
            # Get cash flow data
            cash_flow_data = self._get_cash_flow_data()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Cash Flow Statement"
            
            # Title
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "CASH FLOW STATEMENT"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Date range
            ws.merge_cells('A2:C2')
            date_cell = ws['A2']
            date_cell.value = f"For the period ending {datetime.now().strftime('%B %d, %Y')}"
            date_cell.alignment = Alignment(horizontal='center')
            
            ws.append([''] * 3)  # Empty row
            
            # OPERATING ACTIVITIES
            ws.append(['CASH FLOWS FROM OPERATING ACTIVITIES', '', ''])
            operating_row = ws.max_row
            ws.cell(row=operating_row, column=1).font = Font(bold=True)
            
            total_operating = 0
            for item in cash_flow_data['operating']:
                ws.append([f"  {item['name']}", '', item['amount']])
                total_operating += item['amount']
            
            ws.append(['Net Cash from Operating Activities', '', total_operating])
            total_operating_row = ws.max_row
            ws.cell(row=total_operating_row, column=1).font = Font(bold=True)
            ws.cell(row=total_operating_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # INVESTING ACTIVITIES
            ws.append(['CASH FLOWS FROM INVESTING ACTIVITIES', '', ''])
            investing_row = ws.max_row
            ws.cell(row=investing_row, column=1).font = Font(bold=True)
            
            total_investing = 0
            for item in cash_flow_data['investing']:
                ws.append([f"  {item['name']}", '', item['amount']])
                total_investing += item['amount']
            
            ws.append(['Net Cash from Investing Activities', '', total_investing])
            total_investing_row = ws.max_row
            ws.cell(row=total_investing_row, column=1).font = Font(bold=True)
            ws.cell(row=total_investing_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # FINANCING ACTIVITIES
            ws.append(['CASH FLOWS FROM FINANCING ACTIVITIES', '', ''])
            financing_row = ws.max_row
            ws.cell(row=financing_row, column=1).font = Font(bold=True)
            
            total_financing = 0
            for item in cash_flow_data['financing']:
                ws.append([f"  {item['name']}", '', item['amount']])
                total_financing += item['amount']
            
            ws.append(['Net Cash from Financing Activities', '', total_financing])
            total_financing_row = ws.max_row
            ws.cell(row=total_financing_row, column=1).font = Font(bold=True)
            ws.cell(row=total_financing_row, column=3).font = Font(bold=True)
            
            ws.append([''] * 3)  # Empty row
            
            # NET CHANGE IN CASH
            net_change = total_operating + total_investing + total_financing
            ws.append(['NET CHANGE IN CASH', '', net_change])
            net_change_row = ws.max_row
            ws.cell(row=net_change_row, column=1).font = Font(bold=True)
            ws.cell(row=net_change_row, column=3).font = Font(bold=True)
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"cash_flow_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='cash_flow',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'cash_flow',
                'filename': filename,
                'filepath': filepath,
                'operating_cash_flow': total_operating,
                'investing_cash_flow': total_investing,
                'financing_cash_flow': total_financing,
                'net_change_in_cash': net_change,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating cash flow report: {str(e)}")
            raise
    
    def generate_ledger_summary_report(self, file_id: int) -> Dict[str, Any]:
        """Generate Ledger Summary Report"""
        try:
            # Get ledger summary data
            ledger_data = self._get_ledger_summary_data()
            
            # Create Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Ledger Summary"
            
            # Title
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = "LEDGER SUMMARY"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Date range
            ws.merge_cells('A2:E2')
            date_cell = ws['A2']
            date_cell.value = f"As of {datetime.now().strftime('%B %d, %Y')}"
            date_cell.alignment = Alignment(horizontal='center')
            
            ws.append([''] * 5)  # Empty row
            
            # Headers
            headers = ['Account Code', 'Account Name', 'Opening Balance', 'Closing Balance', 'Net Change']
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Add ledger data
            for account in ledger_data['accounts']:
                row_data = [
                    account['account_code'],
                    account['account_name'],
                    account['opening_balance'],
                    account['closing_balance'],
                    account['net_change']
                ]
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"ledger_summary_{timestamp}.xlsx"
            filepath = os.path.join('reports', filename)
            wb.save(filepath)
            
            # Save report record
            report = FinancialReport(
                report_type='ledger_summary',
                report_period=datetime.now().strftime('%m-%Y'),
                generated_by=1,  # TODO: Get from current user
                file_path=filepath,
                parameters=f'{{"file_id": {file_id}}}'
            )
            db.session.add(report)
            db.session.commit()
            
            return {
                'type': 'ledger_summary',
                'filename': filename,
                'filepath': filepath,
                'total_accounts': len(ledger_data['accounts']),
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating ledger summary report: {str(e)}")
            raise
    
    def _get_balance_sheet_data(self) -> Dict[str, Any]:
        """Get data for balance sheet"""
        # Sample data structure - in real implementation, this would query the database
        return {
            'current_assets': [
                {'name': 'Cash and Cash Equivalents', 'amount': 50000},
                {'name': 'Accounts Receivable', 'amount': 25000},
                {'name': 'Inventory', 'amount': 15000},
                {'name': 'Prepaid Expenses', 'amount': 5000}
            ],
            'fixed_assets': [
                {'name': 'Property, Plant & Equipment', 'amount': 100000},
                {'name': 'Less: Accumulated Depreciation', 'amount': -20000}
            ],
            'current_liabilities': [
                {'name': 'Accounts Payable', 'amount': 15000},
                {'name': 'Accrued Expenses', 'amount': 8000},
                {'name': 'Short-term Debt', 'amount': 10000}
            ],
            'long_term_liabilities': [
                {'name': 'Long-term Debt', 'amount': 50000}
            ],
            'equity': [
                {'name': 'Common Stock', 'amount': 50000},
                {'name': 'Retained Earnings', 'amount': 37000}
            ]
        }
    
    def _get_income_statement_data(self) -> Dict[str, Any]:
        """Get data for income statement"""
        # Sample data structure - in real implementation, this would query the database
        return {
            'revenue': [
                {'name': 'Sales Revenue', 'amount': 200000},
                {'name': 'Service Revenue', 'amount': 50000}
            ],
            'cogs': [
                {'name': 'Cost of Goods Sold', 'amount': 120000}
            ],
            'expenses': [
                {'name': 'Salaries and Wages', 'amount': 60000},
                {'name': 'Rent Expense', 'amount': 24000},
                {'name': 'Utilities', 'amount': 12000},
                {'name': 'Marketing', 'amount': 8000},
                {'name': 'Depreciation', 'amount': 10000}
            ]
        }
    
    def _get_cash_flow_data(self) -> Dict[str, Any]:
        """Get data for cash flow statement"""
        # Sample data structure - in real implementation, this would query the database
        return {
            'operating': [
                {'name': 'Net Income', 'amount': 16000},
                {'name': 'Depreciation', 'amount': 10000},
                {'name': 'Changes in Accounts Receivable', 'amount': -5000},
                {'name': 'Changes in Inventory', 'amount': -2000},
                {'name': 'Changes in Accounts Payable', 'amount': 3000}
            ],
            'investing': [
                {'name': 'Purchase of Equipment', 'amount': -15000},
                {'name': 'Sale of Investments', 'amount': 5000}
            ],
            'financing': [
                {'name': 'Proceeds from Loan', 'amount': 20000},
                {'name': 'Repayment of Debt', 'amount': -10000},
                {'name': 'Dividends Paid', 'amount': -5000}
            ]
        }
    
    def _get_ledger_summary_data(self) -> Dict[str, Any]:
        """Get data for ledger summary"""
        # Sample data structure - in real implementation, this would query the database
        return {
            'accounts': [
                {
                    'account_code': '1001',
                    'account_name': 'Cash',
                    'opening_balance': 45000,
                    'closing_balance': 50000,
                    'net_change': 5000
                },
                {
                    'account_code': '1201',
                    'account_name': 'Accounts Receivable',
                    'opening_balance': 20000,
                    'closing_balance': 25000,
                    'net_change': 5000
                },
                {
                    'account_code': '2001',
                    'account_name': 'Accounts Payable',
                    'opening_balance': 12000,
                    'closing_balance': 15000,
                    'net_change': 3000
                }
            ]
        }
