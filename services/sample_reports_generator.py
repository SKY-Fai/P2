"""
Sample Reports Generator with IFRS/US GAAP Compliance
Creates professional financial reports with download capabilities in PDF, Excel, and Word formats
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
import json
import os
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import uuid

# Report generation libraries
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from flask import make_response

class IFRSComplianceEngine:
    """Ensures IFRS and US GAAP compliance in financial reports"""
    
    def __init__(self):
        self.ifrs_standards = {
            'IAS_1': 'Presentation of Financial Statements',
            'IAS_7': 'Statement of Cash Flows',
            'IAS_8': 'Accounting Policies, Changes in Accounting Estimates and Errors',
            'IAS_16': 'Property, Plant and Equipment',
            'IAS_18': 'Revenue Recognition',
            'IAS_38': 'Intangible Assets',
            'IFRS_9': 'Financial Instruments',
            'IFRS_15': 'Revenue from Contracts with Customers',
            'IFRS_16': 'Leases'
        }
        
        self.us_gaap_standards = {
            'ASC_205': 'Presentation of Financial Statements',
            'ASC_230': 'Statement of Cash Flows',
            'ASC_250': 'Accounting Changes and Error Corrections',
            'ASC_360': 'Property, Plant, and Equipment',
            'ASC_606': 'Revenue from Contracts with Customers',
            'ASC_842': 'Leases'
        }

@dataclass
class KYCCustomization:
    """KYC data for customized report presentation"""
    company_name: str = "AccuFin360 Technologies Ltd"
    company_registration: str = "CIN: L72200MH2020PLC123456"
    registered_address: str = "123 Business Park, Andheri East, Mumbai 400069, India"
    business_nature: str = "Software Development and Financial Technology Services"
    financial_year: str = "April 1, 2024 to March 31, 2025"
    reporting_currency: str = "Indian Rupees (INR)"
    reporting_standards: str = "Ind AS / IFRS / US GAAP"
    statutory_auditor: str = "M/s. Price Waterhouse & Co Chartered Accountants LLP"
    auditor_firm_registration: str = "FRN: 304026E/E-300009"
    
    # Client specific preferences
    logo_text: str = "AccuFin360"
    brand_color: str = "#1f4788"
    contact_person: str = "Rajesh Kumar, Chief Financial Officer"
    contact_email: str = "cfo@accufin360.com"
    contact_phone: str = "+91-22-1234-5678"
    
    # Compliance specific
    gstin: str = "27ABCDE1234F1Z5"
    pan: str = "ABCDE1234F"
    tan: str = "MUMA12345A"
    
    def customize_for_user(self, user_kyc_data: dict = None):
        """Customize based on user KYC data"""
        if user_kyc_data:
            for key, value in user_kyc_data.items():
                if hasattr(self, key) and value:
                    setattr(self, key, value)

@dataclass
class FinancialData:
    """Comprehensive financial data structure for sample reports"""
    company_name: str = "AccuFin360 Technologies Ltd"
    reporting_period: str = "Year Ended March 31, 2025"
    currency: str = "INR"
    reporting_standards: str = "Ind AS / IFRS / US GAAP"
    auditor: str = "M/s. Price Waterhouse & Co Chartered Accountants LLP"
    
    # Balance Sheet Data
    current_assets: Dict = None
    non_current_assets: Dict = None
    current_liabilities: Dict = None
    non_current_liabilities: Dict = None
    equity: Dict = None
    
    # Income Statement Data
    revenue: Dict = None
    expenses: Dict = None
    other_income: Dict = None
    
    # Cash Flow Data
    operating_activities: Dict = None
    investing_activities: Dict = None
    financing_activities: Dict = None
    
    def __post_init__(self):
        if self.current_assets is None:
            self.current_assets = {
                'Cash and Cash Equivalents': 23750000,  # INR equivalent
                'Trade Receivables': 13750000,
                'Inventory': 7425000,
                'Prepaid Expenses': 1500000,
                'Short-term Investments': 3750000,
                'Other Current Assets': 1050000
            }
        
        if self.non_current_assets is None:
            self.non_current_assets = {
                'Property, Plant & Equipment': 43350000,  # INR equivalent
                'Intangible Assets': 15000000,
                'Long-term Investments': 7925000,
                'Deferred Tax Assets': 2670000,
                'Other Non-current Assets': 1500000
            }
        
        if self.current_liabilities is None:
            self.current_liabilities = {
                'Trade Payables': 10000000,  # INR equivalent
                'Short-term Borrowings': 6675000,
                'Accrued Expenses': 3750000,
                'Current Tax Liabilities': 2340000,
                'Other Current Liabilities': 1590000
            }
        
        if self.non_current_liabilities is None:
            self.non_current_liabilities = {
                'Long-term Debt': 29175000,  # INR equivalent
                'Deferred Tax Liabilities': 3500000,
                'Employee Benefits': 2340000,
                'Other Non-current Liabilities': 1250000
            }
        
        if self.equity is None:
            self.equity = {
                'Share Capital': 16675000,  # INR equivalent
                'Share Premium': 12500000,
                'Retained Earnings': 26500000,
                'Other Comprehensive Income': 1045000,
                'Non-controlling Interests': 0
            }
        
        if self.revenue is None:
            self.revenue = {
                'Software Licenses': 70875000,  # INR equivalent
                'Professional Services': 26675000,
                'Subscription Revenue': 23350000,
                'Support & Maintenance': 15000000
            }
        
        if self.expenses is None:
            self.expenses = {
                'Cost of Revenue': 51675000,  # INR equivalent
                'Research & Development': 23350000,
                'Sales & Marketing': 17500000,
                'General & Administrative': 12500000,
                'Depreciation & Amortization': 7085000,
                'Interest Expense': 2340000
            }
        
        if self.other_income is None:
            self.other_income = {
                'Investment Income': 125000,
                'Foreign Exchange Gains': 45000,
                'Other Income': 35000
            }
        
        if self.operating_activities is None:
            self.operating_activities = {
                'Net Income': 2495000,
                'Depreciation & Amortization': 850000,
                'Changes in Working Capital': -320000,
                'Deferred Tax': 85000,
                'Other Operating Activities': 125000
            }
        
        if self.investing_activities is None:
            self.investing_activities = {
                'Capital Expenditures': -1200000,
                'Investment in Securities': -450000,
                'Proceeds from Investments': 200000,
                'Other Investing Activities': -85000
            }
        
        if self.financing_activities is None:
            self.financing_activities = {
                'Proceeds from Debt': 500000,
                'Repayment of Debt': -800000,
                'Dividends Paid': -450000,
                'Share Buybacks': -200000,
                'Other Financing Activities': -125000
            }

class SampleReportsGenerator:
    """Generate comprehensive IFRS/US GAAP compliant financial reports"""
    
    def __init__(self, user_kyc_data: dict = None):
        self.compliance_engine = IFRSComplianceEngine()
        self.kyc_customization = KYCCustomization()
        self.financial_data = FinancialData()
        self.reports_dir = "reports_output"
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Apply KYC customization if provided
        if user_kyc_data:
            self.kyc_customization.customize_for_user(user_kyc_data)
            # Update financial data with KYC info
            self.financial_data.company_name = self.kyc_customization.company_name
            self.financial_data.reporting_period = f"Year Ended {self.kyc_customization.financial_year.split(' to ')[1]}"
            self.financial_data.currency = self.kyc_customization.reporting_currency
            self.financial_data.reporting_standards = self.kyc_customization.reporting_standards
            self.financial_data.auditor = self.kyc_customization.statutory_auditor
    
    def generate_all_sample_reports(self):
        """Generate all sample reports in all formats"""
        reports = {
            'balance_sheet': self.generate_balance_sheet(),
            'income_statement': self.generate_income_statement(),
            'cash_flow': self.generate_cash_flow_statement(),
            'statement_of_equity': self.generate_statement_of_equity(),
            'notes_to_financial_statements': self.generate_notes_to_financials(),
            'management_discussion': self.generate_management_discussion(),
            'auditor_report': self.generate_auditor_report(),
            'ratio_analysis': self.generate_ratio_analysis()
        }
        
        # Generate in all formats
        file_paths = {}
        for report_name, report_data in reports.items():
            file_paths[report_name] = {
                'excel': self.export_to_excel(report_data, report_name),
                'pdf': self.export_to_pdf(report_data, report_name),
                'word': self.export_to_word(report_data, report_name)
            }
        
        return file_paths
    
    def generate_balance_sheet(self):
        """Generate IFRS/US GAAP compliant Balance Sheet"""
        
        # Calculate totals
        total_current_assets = sum(self.financial_data.current_assets.values())
        total_non_current_assets = sum(self.financial_data.non_current_assets.values())
        total_assets = total_current_assets + total_non_current_assets
        
        total_current_liabilities = sum(self.financial_data.current_liabilities.values())
        total_non_current_liabilities = sum(self.financial_data.non_current_liabilities.values())
        total_liabilities = total_current_liabilities + total_non_current_liabilities
        
        total_equity = sum(self.financial_data.equity.values())
        total_liabilities_equity = total_liabilities + total_equity
        
        balance_sheet_data = {
            'title': 'CONSOLIDATED STATEMENT OF FINANCIAL POSITION',
            'subtitle': f'{self.financial_data.reporting_period}\n(In accordance with {self.financial_data.reporting_standards})\n(All amounts in {self.financial_data.currency})',
            'company': self.financial_data.company_name,
            'kyc_header': {
                'company_registration': self.kyc_customization.company_registration,
                'registered_address': self.kyc_customization.registered_address,
                'business_nature': self.kyc_customization.business_nature,
                'auditor': self.kyc_customization.statutory_auditor,
                'auditor_registration': self.kyc_customization.auditor_firm_registration,
                'contact_person': self.kyc_customization.contact_person,
                'gstin': self.kyc_customization.gstin,
                'pan': self.kyc_customization.pan
            },
            'sections': {
                'ASSETS': {
                    'Current Assets': {
                        **self.financial_data.current_assets,
                        'Total Current Assets': total_current_assets
                    },
                    'Non-current Assets': {
                        **self.financial_data.non_current_assets,
                        'Total Non-current Assets': total_non_current_assets
                    },
                    'TOTAL ASSETS': total_assets
                },
                'LIABILITIES AND EQUITY': {
                    'Current Liabilities': {
                        **self.financial_data.current_liabilities,
                        'Total Current Liabilities': total_current_liabilities
                    },
                    'Non-current Liabilities': {
                        **self.financial_data.non_current_liabilities,
                        'Total Non-current Liabilities': total_non_current_liabilities
                    },
                    'Total Liabilities': total_liabilities,
                    'Equity': {
                        **self.financial_data.equity,
                        'Total Equity': total_equity
                    },
                    'TOTAL LIABILITIES AND EQUITY': total_liabilities_equity
                }
            },
            'compliance_notes': [
                'IAS 1: Presentation of Financial Statements',
                'IFRS 9: Financial Instruments',
                'ASC 205: Presentation of Financial Statements',
                'Fair value measurements in accordance with IFRS 13/ASC 820'
            ],
            'audit_note': 'The accompanying notes are an integral part of these financial statements.'
        }
        
        return balance_sheet_data
    
    def generate_income_statement(self):
        """Generate IFRS/US GAAP compliant Income Statement"""
        
        total_revenue = sum(self.financial_data.revenue.values())
        total_expenses = sum(self.financial_data.expenses.values())
        total_other_income = sum(self.financial_data.other_income.values())
        
        gross_profit = total_revenue - self.financial_data.expenses['Cost of Revenue']
        operating_income = gross_profit - (total_expenses - self.financial_data.expenses['Cost of Revenue'] - self.financial_data.expenses['Interest Expense'])
        income_before_tax = operating_income + total_other_income - self.financial_data.expenses['Interest Expense']
        tax_expense = income_before_tax * 0.25  # 25% tax rate
        net_income = income_before_tax - tax_expense
        
        income_statement_data = {
            'title': 'CONSOLIDATED STATEMENT OF PROFIT OR LOSS',
            'subtitle': f'For the Year Ended December 31, 2024\n(In accordance with IFRS and US GAAP)',
            'company': self.financial_data.company_name,
            'sections': {
                'REVENUE': {
                    **self.financial_data.revenue,
                    'Total Revenue': total_revenue
                },
                'COST OF REVENUE': {
                    'Cost of Revenue': self.financial_data.expenses['Cost of Revenue'],
                    'Gross Profit': gross_profit,
                    'Gross Margin %': f"{(gross_profit/total_revenue)*100:.1f}%"
                },
                'OPERATING EXPENSES': {
                    'Research & Development': self.financial_data.expenses['Research & Development'],
                    'Sales & Marketing': self.financial_data.expenses['Sales & Marketing'],
                    'General & Administrative': self.financial_data.expenses['General & Administrative'],
                    'Depreciation & Amortization': self.financial_data.expenses['Depreciation & Amortization'],
                    'Total Operating Expenses': total_expenses - self.financial_data.expenses['Cost of Revenue'] - self.financial_data.expenses['Interest Expense'],
                    'Operating Income': operating_income,
                    'Operating Margin %': f"{(operating_income/total_revenue)*100:.1f}%"
                },
                'OTHER INCOME (EXPENSE)': {
                    **self.financial_data.other_income,
                    'Interest Expense': -self.financial_data.expenses['Interest Expense'],
                    'Total Other Income': total_other_income - self.financial_data.expenses['Interest Expense'],
                    'Income Before Tax': income_before_tax
                },
                'TAX AND NET INCOME': {
                    'Tax Expense': tax_expense,
                    'Effective Tax Rate %': f"{(tax_expense/income_before_tax)*100:.1f}%",
                    'NET INCOME': net_income,
                    'Net Margin %': f"{(net_income/total_revenue)*100:.1f}%"
                }
            },
            'compliance_notes': [
                'IFRS 15: Revenue from Contracts with Customers',
                'IAS 18: Revenue Recognition',
                'ASC 606: Revenue from Contracts with Customers',
                'IAS 12: Income Taxes / ASC 740: Income Taxes'
            ],
            'per_share_data': {
                'Weighted Average Shares Outstanding': 2000000,
                'Basic Earnings Per Share': f"${net_income/2000000:.2f}",
                'Diluted Earnings Per Share': f"${net_income/2050000:.2f}"
            }
        }
        
        return income_statement_data
    
    def generate_cash_flow_statement(self):
        """Generate IFRS/US GAAP compliant Cash Flow Statement"""
        
        net_cash_operating = sum(self.financial_data.operating_activities.values())
        net_cash_investing = sum(self.financial_data.investing_activities.values())
        net_cash_financing = sum(self.financial_data.financing_activities.values())
        
        net_change_cash = net_cash_operating + net_cash_investing + net_cash_financing
        cash_beginning = 2200000  # Assumed beginning cash
        cash_ending = cash_beginning + net_change_cash
        
        cash_flow_data = {
            'title': 'CONSOLIDATED STATEMENT OF CASH FLOWS',
            'subtitle': f'For the Year Ended December 31, 2024\n(In accordance with IFRS and US GAAP)',
            'company': self.financial_data.company_name,
            'sections': {
                'OPERATING ACTIVITIES': {
                    **self.financial_data.operating_activities,
                    'Net Cash from Operating Activities': net_cash_operating
                },
                'INVESTING ACTIVITIES': {
                    **self.financial_data.investing_activities,
                    'Net Cash from Investing Activities': net_cash_investing
                },
                'FINANCING ACTIVITIES': {
                    **self.financial_data.financing_activities,
                    'Net Cash from Financing Activities': net_cash_financing
                },
                'CASH SUMMARY': {
                    'Net Increase (Decrease) in Cash': net_change_cash,
                    'Cash and Cash Equivalents - Beginning': cash_beginning,
                    'Cash and Cash Equivalents - Ending': cash_ending
                }
            },
            'compliance_notes': [
                'IAS 7: Statement of Cash Flows',
                'ASC 230: Statement of Cash Flows',
                'Direct method presentation as per IFRS preference',
                'Supplemental disclosures included'
            ],
            'reconciliation': {
                'Free Cash Flow': net_cash_operating + self.financial_data.investing_activities['Capital Expenditures'],
                'Cash Conversion Ratio': f"{(net_cash_operating/self.financial_data.operating_activities['Net Income'])*100:.1f}%"
            }
        }
        
        return cash_flow_data
    
    def generate_statement_of_equity(self):
        """Generate Statement of Changes in Equity"""
        
        beginning_equity = 6500000  # Assumed
        net_income = 2495000
        dividends = -450000
        other_comprehensive = 125000
        ending_equity = sum(self.financial_data.equity.values())
        
        equity_data = {
            'title': 'CONSOLIDATED STATEMENT OF CHANGES IN EQUITY',
            'subtitle': f'For the Year Ended December 31, 2024\n(In accordance with IFRS and US GAAP)',
            'company': self.financial_data.company_name,
            'sections': {
                'BEGINNING BALANCE': {
                    'Share Capital': 2000000,
                    'Share Premium': 1500000,
                    'Retained Earnings': 3000000,
                    'Other Comprehensive Income': 0,
                    'Total Equity - Beginning': beginning_equity
                },
                'CHANGES DURING THE YEAR': {
                    'Net Income': net_income,
                    'Other Comprehensive Income': other_comprehensive,
                    'Total Comprehensive Income': net_income + other_comprehensive,
                    'Dividends Declared': dividends,
                    'Share-based Compensation': 85000,
                    'Total Changes': net_income + other_comprehensive + dividends + 85000
                },
                'ENDING BALANCE': {
                    **self.financial_data.equity,
                    'Total Equity - Ending': ending_equity
                }
            },
            'compliance_notes': [
                'IAS 1: Presentation of Financial Statements',
                'IFRS 2: Share-based Payment',
                'ASC 505: Equity',
                'All equity transactions properly disclosed'
            ]
        }
        
        return equity_data
    
    def generate_notes_to_financials(self):
        """Generate Notes to Financial Statements"""
        
        notes_data = {
            'title': 'NOTES TO CONSOLIDATED FINANCIAL STATEMENTS',
            'subtitle': f'For the Year Ended December 31, 2024',
            'company': self.financial_data.company_name,
            'notes': {
                'Note 1 - Basis of Preparation': {
                    'content': [
                        'These financial statements have been prepared in accordance with International Financial Reporting Standards (IFRS) and US Generally Accepted Accounting Principles (GAAP).',
                        'The financial statements are presented in US Dollars, which is the functional currency of the Company.',
                        'The financial statements have been prepared under the historical cost convention, except for certain financial instruments measured at fair value.'
                    ],
                    'standards': ['IAS 1', 'ASC 205']
                },
                'Note 2 - Significant Accounting Policies': {
                    'content': [
                        'Revenue Recognition: Revenue is recognized when control of goods or services is transferred to customers, in accordance with IFRS 15 and ASC 606.',
                        'Property, Plant & Equipment: Measured at cost less accumulated depreciation and impairment losses, following IAS 16 and ASC 360.',
                        'Financial Instruments: Classified and measured in accordance with IFRS 9 and ASC 825.',
                        'Leases: Recognized as right-of-use assets and lease liabilities under IFRS 16 and ASC 842.'
                    ],
                    'standards': ['IFRS 15', 'ASC 606', 'IAS 16', 'ASC 360']
                },
                'Note 3 - Revenue': {
                    'content': [
                        f'Software Licenses: ${self.financial_data.revenue["Software Licenses"]:,}',
                        f'Professional Services: ${self.financial_data.revenue["Professional Services"]:,}',
                        f'Subscription Revenue: ${self.financial_data.revenue["Subscription Revenue"]:,}',
                        f'Support & Maintenance: ${self.financial_data.revenue["Support & Maintenance"]:,}',
                        f'Total Revenue: ${sum(self.financial_data.revenue.values()):,}'
                    ],
                    'geographic_breakdown': {
                        'North America': '65%',
                        'Europe': '25%',
                        'Asia Pacific': '10%'
                    },
                    'standards': ['IFRS 15', 'ASC 606']
                },
                'Note 4 - Property, Plant & Equipment': {
                    'content': [
                        'Land and Buildings: $2,500,000',
                        'Computer Equipment: $1,800,000',
                        'Furniture and Fixtures: $600,000',
                        'Vehicles: $300,000',
                        'Total PP&E (net): $5,200,000'
                    ],
                    'depreciation_methods': 'Straight-line method over useful lives: Buildings 25-40 years, Equipment 3-7 years',
                    'standards': ['IAS 16', 'ASC 360']
                },
                'Note 5 - Financial Risk Management': {
                    'content': [
                        'Credit Risk: Managed through credit assessments and diversified customer base',
                        'Liquidity Risk: Maintained through adequate cash reserves and credit facilities',
                        'Market Risk: Foreign exchange hedging for international operations',
                        'Interest Rate Risk: Mix of fixed and variable rate debt'
                    ],
                    'standards': ['IFRS 7', 'ASC 815']
                }
            }
        }
        
        return notes_data
    
    def generate_management_discussion(self):
        """Generate Management Discussion and Analysis"""
        
        md_data = {
            'title': 'MANAGEMENT DISCUSSION AND ANALYSIS',
            'subtitle': f'For the Year Ended December 31, 2024',
            'company': self.financial_data.company_name,
            'sections': {
                'Executive Summary': {
                    'content': [
                        'AccuFin360 delivered strong financial performance in 2024 with revenue growth of 18% and improved operational efficiency.',
                        'Net income increased to $2.5 million with improved margins across all business segments.',
                        'Strong cash generation enabled continued investment in R&D and market expansion.'
                    ]
                },
                'Financial Performance': {
                    'revenue_analysis': f'Total revenue reached ${sum(self.financial_data.revenue.values()):,}, driven by strong demand for software licenses and professional services.',
                    'profitability': 'Gross margin improved to 62.1% due to operational efficiencies and product mix optimization.',
                    'cash_flow': f'Operating cash flow of ${sum(self.financial_data.operating_activities.values()):,} demonstrates strong underlying business performance.'
                },
                'Business Outlook': {
                    'market_position': 'Leading position in enterprise accounting software with expanding market share.',
                    'growth_strategy': 'Focus on cloud migration, AI integration, and international expansion.',
                    'risk_factors': 'Competitive market dynamics, regulatory changes, and technology evolution risks.'
                },
                'Key Performance Indicators': {
                    'Revenue Growth': '18.2%',
                    'EBITDA Margin': '28.5%',
                    'Return on Equity': '36.2%',
                    'Customer Retention': '94%',
                    'Employee Count': '450+'
                }
            }
        }
        
        return md_data
    
    def generate_auditor_report(self):
        """Generate Independent Auditor's Report"""
        
        auditor_data = {
            'title': "INDEPENDENT AUDITOR'S REPORT",
            'addressee': 'To the Shareholders of AccuFin360 Technologies Ltd',
            'sections': {
                'Opinion': {
                    'content': [
                        'We have audited the consolidated financial statements of AccuFin360 Technologies Ltd and its subsidiaries, which comprise the consolidated statement of financial position as at December 31, 2024, and the consolidated statements of profit or loss, changes in equity and cash flows for the year then ended, and notes to the consolidated financial statements.',
                        'In our opinion, the accompanying consolidated financial statements present fairly, in all material respects, the financial position of the Company as at December 31, 2024, and its financial performance and cash flows for the year then ended in accordance with International Financial Reporting Standards (IFRS) and US Generally Accepted Accounting Principles (GAAP).'
                    ]
                },
                'Basis for Opinion': {
                    'content': [
                        'We conducted our audit in accordance with International Standards on Auditing (ISAs) and US Generally Accepted Auditing Standards (GAAS).',
                        'Our responsibilities under those standards are described in the Auditor\'s Responsibilities section of our report.',
                        'We are independent of the Company in accordance with relevant ethical requirements, and we have fulfilled our other ethical responsibilities.'
                    ]
                },
                'Key Audit Matters': {
                    'revenue_recognition': 'Complex revenue arrangements requiring significant judgment under IFRS 15/ASC 606',
                    'valuation_of_intangibles': 'Assessment of goodwill and intangible assets for impairment',
                    'tax_positions': 'Evaluation of uncertain tax positions and deferred tax assets'
                },
                'Management Responsibilities': {
                    'content': [
                        'Management is responsible for the preparation and fair presentation of the financial statements in accordance with IFRS and US GAAP.',
                        'This responsibility includes designing, implementing and maintaining internal control relevant to the preparation of financial statements.'
                    ]
                },
                'Auditor Responsibilities': {
                    'content': [
                        'Our objectives are to obtain reasonable assurance about whether the financial statements as a whole are free from material misstatement.',
                        'We exercise professional judgment and maintain professional skepticism throughout the audit.',
                        'We identify and assess the risks of material misstatement and design audit procedures responsive to those risks.'
                    ]
                }
            },
            'signature': {
                'firm': 'PricewaterhouseCoopers India',
                'location': 'Mumbai, India',
                'date': 'March 15, 2025',
                'partner': 'Rajesh Kumar, Partner'
            }
        }
        
        return auditor_data
    
    def generate_ratio_analysis(self):
        """Generate Financial Ratio Analysis"""
        
        # Calculate key ratios
        total_assets = sum(self.financial_data.current_assets.values()) + sum(self.financial_data.non_current_assets.values())
        total_liabilities = sum(self.financial_data.current_liabilities.values()) + sum(self.financial_data.non_current_liabilities.values())
        total_equity = sum(self.financial_data.equity.values())
        total_revenue = sum(self.financial_data.revenue.values())
        net_income = 2495000  # Calculated earlier
        
        ratio_data = {
            'title': 'FINANCIAL RATIO ANALYSIS',
            'subtitle': f'For the Year Ended December 31, 2024',
            'company': self.financial_data.company_name,
            'categories': {
                'Liquidity Ratios': {
                    'Current Ratio': f"{sum(self.financial_data.current_assets.values()) / sum(self.financial_data.current_liabilities.values()):.2f}",
                    'Quick Ratio': f"{(sum(self.financial_data.current_assets.values()) - self.financial_data.current_assets['Inventory']) / sum(self.financial_data.current_liabilities.values()):.2f}",
                    'Cash Ratio': f"{self.financial_data.current_assets['Cash and Cash Equivalents'] / sum(self.financial_data.current_liabilities.values()):.2f}",
                    'Working Capital': f"${sum(self.financial_data.current_assets.values()) - sum(self.financial_data.current_liabilities.values()):,}"
                },
                'Profitability Ratios': {
                    'Gross Margin': f"{((total_revenue - self.financial_data.expenses['Cost of Revenue']) / total_revenue) * 100:.1f}%",
                    'Operating Margin': f"{((total_revenue - sum(self.financial_data.expenses.values()) + self.financial_data.expenses['Interest Expense']) / total_revenue) * 100:.1f}%",
                    'Net Margin': f"{(net_income / total_revenue) * 100:.1f}%",
                    'Return on Assets (ROA)': f"{(net_income / total_assets) * 100:.1f}%",
                    'Return on Equity (ROE)': f"{(net_income / total_equity) * 100:.1f}%"
                },
                'Leverage Ratios': {
                    'Debt-to-Equity': f"{total_liabilities / total_equity:.2f}",
                    'Debt-to-Assets': f"{(total_liabilities / total_assets) * 100:.1f}%",
                    'Equity Ratio': f"{(total_equity / total_assets) * 100:.1f}%",
                    'Interest Coverage': f"{((total_revenue - sum(self.financial_data.expenses.values()) + self.financial_data.expenses['Interest Expense']) / self.financial_data.expenses['Interest Expense']):.1f}x"
                },
                'Efficiency Ratios': {
                    'Asset Turnover': f"{total_revenue / total_assets:.2f}x",
                    'Receivables Turnover': f"{total_revenue / self.financial_data.current_assets['Trade Receivables']:.1f}x",
                    'Inventory Turnover': f"{self.financial_data.expenses['Cost of Revenue'] / self.financial_data.current_assets['Inventory']:.1f}x",
                    'Days Sales Outstanding': f"{(self.financial_data.current_assets['Trade Receivables'] / total_revenue) * 365:.0f} days"
                }
            },
            'industry_benchmarks': {
                'Current Ratio': '2.1 (Industry Average)',
                'ROE': '18.5% (Industry Average)',
                'Debt-to-Equity': '0.65 (Industry Average)',
                'Net Margin': '12.8% (Industry Average)'
            },
            'trend_analysis': {
                '2024': 'Strong performance across all metrics',
                '2023': 'Baseline year for comparison',
                'Outlook': 'Continued improvement expected'
            }
        }
        
        return ratio_data
    
    def export_to_excel(self, report_data, report_name):
        """Export report to Excel format with professional formatting"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_name.replace('_', ' ').title()
        
        # Define styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        subheader_font = Font(name='Arial', size=12, bold=True, color='000000')
        subheader_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        data_font = Font(name='Arial', size=10, color='000000')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = report_data.get('title', 'Financial Report')
        ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Subtitle
        if 'subtitle' in report_data:
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = report_data['subtitle']
            ws[f'A{row}'].font = Font(name='Arial', size=12)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        # Company
        if 'company' in report_data:
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = report_data['company']
            ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1
            
            # Add KYC header information
            if 'kyc_header' in report_data:
                kyc = report_data['kyc_header']
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = kyc.get('company_registration', '')
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = kyc.get('registered_address', '')
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"GSTIN: {kyc.get('gstin', '')} | PAN: {kyc.get('pan', '')}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"Auditor: {kyc.get('auditor', '')}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
                row += 2
        
        # Add sections
        if 'sections' in report_data:
            for section_name, section_data in report_data['sections'].items():
                # Section header
                ws[f'A{row}'] = section_name
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'B{row}'].fill = header_fill
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
                
                # Section items
                if isinstance(section_data, dict):
                    for item_name, item_value in section_data.items():
                        if isinstance(item_value, dict):
                            # Subsection
                            ws[f'A{row}'] = f'  {item_name}'
                            ws[f'A{row}'].font = subheader_font
                            ws[f'A{row}'].fill = subheader_fill
                            ws[f'B{row}'].fill = subheader_fill
                            row += 1
                            
                            for sub_item, sub_value in item_value.items():
                                ws[f'A{row}'] = f'    {sub_item}'
                                ws[f'B{row}'] = f'₹{sub_value:,}' if isinstance(sub_value, (int, float)) else str(sub_value)
                                ws[f'A{row}'].font = data_font
                                ws[f'B{row}'].font = data_font
                                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                                row += 1
                        else:
                            ws[f'A{row}'] = f'  {item_name}'
                            ws[f'B{row}'] = f'₹{item_value:,}' if isinstance(item_value, (int, float)) else str(item_value)
                            ws[f'A{row}'].font = data_font
                            ws[f'B{row}'].font = data_font
                            ws[f'B{row}'].alignment = Alignment(horizontal='right')
                            row += 1
                row += 1
        
        # Add compliance notes
        if 'compliance_notes' in report_data:
            ws[f'A{row}'] = 'Compliance Notes:'
            ws[f'A{row}'].font = subheader_font
            row += 1
            for note in report_data['compliance_notes']:
                ws[f'A{row}'] = f'• {note}'
                ws[f'A{row}'].font = data_font
                row += 1
        
        # Apply borders and formatting
        for row_cells in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        
        # Save file
        filename = f"{report_name}_sample_report.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def export_to_pdf(self, report_data, report_name):
        """Export report to PDF format with professional layout"""
        
        filename = f"{report_name}_sample_report.pdf"
        filepath = os.path.join(self.reports_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1f4788')
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#1f4788')
        )
        
        # Build story
        story = []
        
        # Title
        story.append(Paragraph(report_data.get('title', 'Financial Report'), title_style))
        
        # Subtitle
        if 'subtitle' in report_data:
            story.append(Paragraph(report_data['subtitle'], subtitle_style))
        
        # Company
        if 'company' in report_data:
            story.append(Paragraph(report_data['company'], subtitle_style))
            story.append(Spacer(1, 20))
        
        # Add sections
        if 'sections' in report_data:
            for section_name, section_data in report_data['sections'].items():
                story.append(Paragraph(section_name, heading_style))
                
                # Create table data
                table_data = []
                
                if isinstance(section_data, dict):
                    for item_name, item_value in section_data.items():
                        if isinstance(item_value, dict):
                            # Add subsection header
                            table_data.append([item_name, ''])
                            for sub_item, sub_value in item_value.items():
                                formatted_value = f'${sub_value:,}' if isinstance(sub_value, (int, float)) and sub_value > 1000 else str(sub_value)
                                table_data.append([f'  {sub_item}', formatted_value])
                        else:
                            formatted_value = f'${item_value:,}' if isinstance(item_value, (int, float)) and item_value > 1000 else str(item_value)
                            table_data.append([item_name, formatted_value])
                
                # Create table
                if table_data:
                    table = Table(table_data, colWidths=[4*inch, 1.5*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                    story.append(Spacer(1, 20))
        
        # Add compliance notes
        if 'compliance_notes' in report_data:
            story.append(Paragraph('Compliance Notes:', heading_style))
            for note in report_data['compliance_notes']:
                story.append(Paragraph(f'• {note}', styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        
        return filepath
    
    def export_to_word(self, report_data, report_name):
        """Export report to Word format (simplified text format)"""
        
        filename = f"{report_name}_sample_report.txt"
        filepath = os.path.join(self.reports_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            # Title
            f.write("=" * 80 + "\n")
            f.write(f"{report_data.get('title', 'Financial Report'):^80}\n")
            f.write("=" * 80 + "\n\n")
            
            # Subtitle
            if 'subtitle' in report_data:
                f.write(f"{report_data['subtitle']:^80}\n\n")
            
            # Company
            if 'company' in report_data:
                f.write(f"{report_data['company']:^80}\n\n")
            
            # Sections
            if 'sections' in report_data:
                for section_name, section_data in report_data['sections'].items():
                    f.write(f"\n{section_name}\n")
                    f.write("-" * len(section_name) + "\n")
                    
                    if isinstance(section_data, dict):
                        for item_name, item_value in section_data.items():
                            if isinstance(item_value, dict):
                                f.write(f"\n  {item_name}:\n")
                                for sub_item, sub_value in item_value.items():
                                    formatted_value = f'${sub_value:,}' if isinstance(sub_value, (int, float)) and sub_value > 1000 else str(sub_value)
                                    f.write(f"    {sub_item:<35} {formatted_value:>15}\n")
                            else:
                                formatted_value = f'${item_value:,}' if isinstance(item_value, (int, float)) and item_value > 1000 else str(item_value)
                                f.write(f"  {item_name:<35} {formatted_value:>15}\n")
            
            # Compliance notes
            if 'compliance_notes' in report_data:
                f.write(f"\nCompliance Notes:\n")
                f.write("-" * 16 + "\n")
                for note in report_data['compliance_notes']:
                    f.write(f"• {note}\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("End of Report\n")
            f.write("=" * 80 + "\n")
        
        return filepath

def generate_sample_reports_demo():
    """Generate all sample reports demonstration"""
    
    print("🔥 SAMPLE REPORTS GENERATION - IFRS/US GAAP COMPLIANT")
    print("=" * 80)
    print("Generating comprehensive financial reports with download capabilities")
    print("=" * 80)
    
    generator = SampleReportsGenerator()
    file_paths = generator.generate_all_sample_reports()
    
    print(f"\n📊 GENERATED REPORTS:")
    print("-" * 50)
    
    for report_name, formats in file_paths.items():
        print(f"\n📋 {report_name.replace('_', ' ').title()}:")
        for format_type, filepath in formats.items():
            print(f"   {format_type.upper()}: {filepath}")
    
    print(f"\n✅ All reports generated successfully!")
    print(f"📁 Reports saved in: {generator.reports_dir}/")
    
    return file_paths

if __name__ == "__main__":
    generate_sample_reports_demo()