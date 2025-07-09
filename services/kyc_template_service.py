"""
KYC Template Service - Creates professional financial report templates with KYC mapping
Generates Excel, PDF, and Word templates with client branding and compliance structure
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import logging

logger = logging.getLogger(__name__)

def create_kyc_mapped_excel_templates(file_path, user):
    """Create comprehensive Excel template with all financial reports and KYC mapping"""
    try:
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create KYC Information Sheet
        create_kyc_info_sheet(wb, user)
        
        # Create report template sheets
        create_journal_template_sheet(wb, user)
        create_ledger_template_sheet(wb, user)
        create_trial_balance_template_sheet(wb, user)
        create_profit_loss_template_sheet(wb, user)
        create_balance_sheet_template_sheet(wb, user)
        create_cash_flow_template_sheet(wb, user)
        create_shareholders_equity_template_sheet(wb, user)
        create_mis_report_template_sheet(wb, user)
        
        # Create instructions sheet
        create_instructions_sheet(wb)
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"KYC-mapped Excel templates created successfully: {file_path}")
        
    except Exception as e:
        logger.error(f"Error creating KYC-mapped Excel templates: {str(e)}")
        raise

def create_kyc_info_sheet(wb, user):
    """Create KYC information mapping sheet"""
    ws = wb.create_sheet("KYC Client Information")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    # KYC Information Header
    ws['A1'] = "CLIENT KYC INFORMATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:D1')
    
    # Client Information Fields
    kyc_fields = [
        ("Client Name", user.get_full_name() if user else "[Client Name]"),
        ("Company Name", "[Company Name]"),
        ("Registration Number", "[Registration Number]"),
        ("Tax ID", "[Tax ID]"),
        ("Address", "[Complete Address]"),
        ("Phone", "[Phone Number]"),
        ("Email", user.email if user else "[Email Address]"),
        ("Financial Year Start", "[MM-DD]"),
        ("Currency", "[Currency Code]"),
        ("Industry", "[Industry Type]"),
        ("Report Period", "[Report Period]"),
        ("Prepared By", user.get_full_name() if user else "[Preparer Name]"),
        ("Date Prepared", datetime.now().strftime('%Y-%m-%d')),
        ("Authorized Signatory", "[Authorized Person]")
    ]
    
    row = 3
    for field, value in kyc_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_journal_template_sheet(wb, user):
    """Create Journal Report template sheet"""
    ws = wb.create_sheet("Journal Report Template")
    
    # Create professional header with KYC mapping
    create_report_header(ws, "JOURNAL REPORT", user)
    
    # Journal entries table headers
    headers = [
        "Date", "Reference", "Account Code", "Account Name", 
        "Description", "Debit Amount", "Credit Amount"
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    sample_data = [
        ["2024-01-01", "JE001", "1001", "Cash Account", "Opening Balance", "50000", ""],
        ["2024-01-01", "JE001", "3001", "Equity Account", "Opening Balance", "", "50000"],
        ["[Date]", "[Reference]", "[Account Code]", "[Account Name]", "[Description]", "[Debit]", "[Credit]"]
    ]
    
    for row_idx, data in enumerate(sample_data, 7):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def create_ledger_template_sheet(wb, user):
    """Create Ledger Report template sheet"""
    ws = wb.create_sheet("Ledger Report Template")
    
    create_report_header(ws, "LEDGER REPORT", user)
    
    # Ledger headers
    headers = [
        "Account Code", "Account Name", "Date", "Reference", 
        "Description", "Debit", "Credit", "Balance"
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_trial_balance_template_sheet(wb, user):
    """Create Trial Balance template sheet"""
    ws = wb.create_sheet("Trial Balance Template")
    
    create_report_header(ws, "TRIAL BALANCE", user)
    
    headers = ["Account Code", "Account Name", "Debit Balance", "Credit Balance"]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_profit_loss_template_sheet(wb, user):
    """Create Profit & Loss Statement template sheet"""
    ws = wb.create_sheet("P&L Statement Template")
    
    create_report_header(ws, "PROFIT & LOSS STATEMENT", user)
    
    # P&L structure
    pnl_items = [
        ("REVENUE", ""),
        ("Sales Revenue", "[Amount]"),
        ("Other Income", "[Amount]"),
        ("Total Revenue", "[Total]"),
        ("", ""),
        ("EXPENSES", ""),
        ("Cost of Goods Sold", "[Amount]"),
        ("Operating Expenses", "[Amount]"),
        ("Administrative Expenses", "[Amount]"),
        ("Total Expenses", "[Total]"),
        ("", ""),
        ("NET PROFIT/LOSS", "[Amount]")
    ]
    
    row = 6
    for item, amount in pnl_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = amount
        if item in ["REVENUE", "EXPENSES", "NET PROFIT/LOSS"]:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

def create_balance_sheet_template_sheet(wb, user):
    """Create Balance Sheet template sheet"""
    ws = wb.create_sheet("Balance Sheet Template")
    
    create_report_header(ws, "BALANCE SHEET", user)
    
    # Balance Sheet structure
    bs_items = [
        ("ASSETS", ""),
        ("Current Assets", ""),
        ("  Cash and Cash Equivalents", "[Amount]"),
        ("  Accounts Receivable", "[Amount]"),
        ("  Inventory", "[Amount]"),
        ("Total Current Assets", "[Total]"),
        ("", ""),
        ("Non-Current Assets", ""),
        ("  Fixed Assets", "[Amount]"),
        ("  Intangible Assets", "[Amount]"),
        ("Total Non-Current Assets", "[Total]"),
        ("TOTAL ASSETS", "[Total]"),
        ("", ""),
        ("LIABILITIES & EQUITY", ""),
        ("Current Liabilities", ""),
        ("  Accounts Payable", "[Amount]"),
        ("  Short-term Debt", "[Amount]"),
        ("Total Current Liabilities", "[Total]"),
        ("", ""),
        ("Equity", ""),
        ("  Share Capital", "[Amount]"),
        ("  Retained Earnings", "[Amount]"),
        ("Total Equity", "[Total]"),
        ("TOTAL LIABILITIES & EQUITY", "[Total]")
    ]
    
    row = 6
    for item, amount in bs_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = amount
        if item in ["ASSETS", "LIABILITIES & EQUITY"] or "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

def create_cash_flow_template_sheet(wb, user):
    """Create Cash Flow Statement template sheet"""
    ws = wb.create_sheet("Cash Flow Template")
    
    create_report_header(ws, "CASH FLOW STATEMENT", user)
    
    # Cash Flow structure
    cf_items = [
        ("OPERATING ACTIVITIES", ""),
        ("Net Income", "[Amount]"),
        ("Adjustments:", ""),
        ("  Depreciation", "[Amount]"),
        ("  Changes in Working Capital", "[Amount]"),
        ("Net Cash from Operating Activities", "[Total]"),
        ("", ""),
        ("INVESTING ACTIVITIES", ""),
        ("Purchase of Fixed Assets", "[Amount]"),
        ("Sale of Assets", "[Amount]"),
        ("Net Cash from Investing Activities", "[Total]"),
        ("", ""),
        ("FINANCING ACTIVITIES", ""),
        ("Borrowings", "[Amount]"),
        ("Loan Repayments", "[Amount]"),
        ("Dividends Paid", "[Amount]"),
        ("Net Cash from Financing Activities", "[Total]"),
        ("", ""),
        ("NET CHANGE IN CASH", "[Amount]"),
        ("Cash at Beginning of Period", "[Amount]"),
        ("CASH AT END OF PERIOD", "[Total]")
    ]
    
    row = 6
    for item, amount in cf_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = amount
        if "ACTIVITIES" in item or "NET" in item and "CHANGE" in item:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

def create_shareholders_equity_template_sheet(wb, user):
    """Create Shareholders' Equity template sheet"""
    ws = wb.create_sheet("Shareholders Equity Template")
    
    create_report_header(ws, "STATEMENT OF SHAREHOLDERS' EQUITY", user)
    
    headers = [
        "Description", "Share Capital", "Retained Earnings", "Total Equity"
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_mis_report_template_sheet(wb, user):
    """Create MIS Report with Ratio Analysis template sheet"""


class KYCTemplateService:
    """Service for creating individual reports with KYC mapping"""
    
    def __init__(self):
        self.reports_dir = 'reports'
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def create_individual_report_excel(self, report_data, report_type, user_info):
        """Create individual Excel report with KYC mapping"""
        try:
            wb = Workbook()
            ws = wb.active
            
            # Report title mapping
            report_titles = {
                'journal': 'JOURNAL REPORT',
                'ledger': 'LEDGER REPORT', 
                'trial_balance': 'TRIAL BALANCE',
                'profit_loss': 'PROFIT & LOSS STATEMENT',
                'balance_sheet': 'BALANCE SHEET',
                'cash_flow': 'CASH FLOW STATEMENT',
                'shareholders_equity': 'SHAREHOLDERS\' EQUITY STATEMENT',
                'mis_report': 'MIS REPORT WITH RATIO ANALYSIS'
            }
            
            title = report_titles.get(report_type, report_type.upper())
            ws.title = title
            
            # Create header with KYC info
            self._create_report_header(ws, title, user_info)
            
            # Add report data
            start_row = 8
            if isinstance(report_data, dict):
                if 'entries' in report_data:
                    self._add_table_data(ws, report_data['entries'], start_row)
                elif 'data' in report_data:
                    self._add_table_data(ws, report_data['data'], start_row)
                else:
                    self._add_dict_data(ws, report_data, start_row)
            elif isinstance(report_data, list):
                self._add_table_data(ws, report_data, start_row)
            
            # Save file
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Error creating Excel report: {str(e)}")
    
    def create_individual_report_pdf(self, report_data, report_type, user_info):
        """Create individual PDF report with KYC mapping"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = os.path.join(self.reports_dir, filename)
            
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            report_titles = {
                'journal': 'JOURNAL REPORT',
                'ledger': 'LEDGER REPORT',
                'trial_balance': 'TRIAL BALANCE', 
                'profit_loss': 'PROFIT & LOSS STATEMENT',
                'balance_sheet': 'BALANCE SHEET',
                'cash_flow': 'CASH FLOW STATEMENT',
                'shareholders_equity': 'SHAREHOLDERS\' EQUITY STATEMENT',
                'mis_report': 'MIS REPORT WITH RATIO ANALYSIS'
            }
            
            title = report_titles.get(report_type, report_type.upper())
            story.append(Paragraph(title, title_style))
            
            # KYC Information
            story.append(Paragraph(f"Prepared for: {user_info['name']}", styles['Normal']))
            story.append(Paragraph(f"Email: {user_info['email']}", styles['Normal']))
            story.append(Paragraph(f"Company: {user_info['company']}", styles['Normal']))
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Add data as table
            if isinstance(report_data, dict) and 'entries' in report_data:
                data = report_data['entries'][:10]  # Limit for PDF
            elif isinstance(report_data, list):
                data = report_data[:10]  # Limit for PDF
            else:
                data = [["Description", "Amount"], ["Report Data", str(report_data)]]
            
            if data:
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
            
            doc.build(story)
            return file_path
            
        except Exception as e:
            raise Exception(f"Error creating PDF report: {str(e)}")
    
    def create_individual_report_word(self, report_data, report_type, user_info):
        """Create individual Word report with KYC mapping"""
        try:
            # Simple Word document creation (basic implementation)
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.doc"
            file_path = os.path.join(self.reports_dir, filename)
            
            report_titles = {
                'journal': 'JOURNAL REPORT',
                'ledger': 'LEDGER REPORT',
                'trial_balance': 'TRIAL BALANCE',
                'profit_loss': 'PROFIT & LOSS STATEMENT', 
                'balance_sheet': 'BALANCE SHEET',
                'cash_flow': 'CASH FLOW STATEMENT',
                'shareholders_equity': 'SHAREHOLDERS\' EQUITY STATEMENT',
                'mis_report': 'MIS REPORT WITH RATIO ANALYSIS'
            }
            
            title = report_titles.get(report_type, report_type.upper())
            
            # Create basic text document
            content = f"""
{title}

Prepared for: {user_info['name']}
Email: {user_info['email']}
Company: {user_info['company']}
Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

REPORT DATA:
{str(report_data)}
"""
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Error creating Word report: {str(e)}")
    
    def _create_report_header(self, ws, title, user_info):
        """Create professional header with KYC information"""
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # KYC Information
        ws['A3'] = f"Prepared for: {user_info['name']}"
        ws['A4'] = f"Email: {user_info['email']}"
        ws['A5'] = f"Company: {user_info['company']}"
        ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    def _add_table_data(self, ws, data, start_row):
        """Add table data to worksheet"""
        if not data:
            return
        
        # Add headers if data is list of dicts
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header).font = Font(bold=True)
            
            # Add data rows
            for row_idx, item in enumerate(data, start_row + 1):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))
        else:
            # Add data as is
            for row_idx, item in enumerate(data, start_row):
                ws.cell(row=row_idx, column=1, value=str(item))
    
    def _add_dict_data(self, ws, data, start_row):
        """Add dictionary data to worksheet"""
        row = start_row
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

def create_instructions_sheet(wb):
    """Create instructions sheet for template usage"""
    ws = wb.create_sheet("Instructions")
    
    # Instructions header
    ws['A1'] = "FINANCIAL REPORT TEMPLATES - USAGE INSTRUCTIONS"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:C1')
    
    instructions = [
        "",
        "OVERVIEW:",
        "These templates are designed with KYC (Know Your Customer) mapping for automatic",
        "client presentation formatting. Each template includes professional structure",
        "and compliance requirements for financial reporting.",
        "",
        "USAGE INSTRUCTIONS:",
        "1. Start with the 'KYC Client Information' sheet",
        "2. Fill in all client details completely",
        "3. Use individual report templates as needed",
        "4. Replace placeholder values with actual data",
        "5. Maintain professional formatting",
        "",
        "TEMPLATE FEATURES:",
        "• Automated client branding integration",
        "• Professional presentation formatting",
        "• Compliance structure included",
        "• Multi-format export ready",
        "• KYC mapping for personalization",
        "",
        "SUPPORT:",
        "For technical support or customization requirements,",
        "contact your financial reporting administrator.",
        "",
        "© F-AI Accountant - Professional Financial Reporting System"
    ]
    
    row = 3
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        if instruction.endswith(":"):
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

def create_report_header(ws, report_title, user):
    """Create professional report header with KYC mapping"""
    # Company name (mapped from KYC)
    ws['A1'] = f"[Company Name] - {report_title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:G1')
    
    # Report period
    ws['A2'] = "For the period ending: [Report Date]"
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:G2')
    
    # Prepared by (mapped from KYC)
    ws['A3'] = f"Prepared by: {user.get_full_name() if user else '[Preparer Name]'}"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A3:G3')

def create_kyc_mapped_pdf_templates(file_path, user):
    """Create PDF templates with KYC mapping"""
    try:
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        story.append(Paragraph("FINANCIAL REPORT TEMPLATES", title_style))
        story.append(Paragraph("KYC-Mapped Professional Templates", styles['Heading2']))
        story.append(Spacer(1, 20))
        
        # KYC Information Section
        story.append(Paragraph("Client Information Template", styles['Heading3']))
        
        kyc_data = [
            ['Field', 'Value'],
            ['Client Name', user.get_full_name() if user else '[Client Name]'],
            ['Company Name', '[Company Name]'],
            ['Email', user.email if user else '[Email Address]'],
            ['Report Period', '[Report Period]'],
            ['Prepared By', user.get_full_name() if user else '[Preparer Name]'],
            ['Date Prepared', datetime.now().strftime('%Y-%m-%d')]
        ]
        
        kyc_table = Table(kyc_data)
        kyc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(kyc_table)
        story.append(Spacer(1, 30))
        
        # Report Templates List
        story.append(Paragraph("Available Report Templates", styles['Heading3']))
        
        templates_data = [
            ['Report Type', 'Description'],
            ['Journal Report', 'Detailed transaction entries'],
            ['Ledger Report', 'Account-wise summaries'],
            ['Trial Balance', 'Balance verification'],
            ['P&L Statement', 'Revenue and expenses'],
            ['Balance Sheet', 'Assets, liabilities, equity'],
            ['Cash Flow Statement', 'Cash movements analysis'],
            ['Shareholders Equity', 'Equity changes statement'],
            ['MIS Report', 'Management ratios and KPIs']
        ]
        
        templates_table = Table(templates_data)
        templates_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(templates_table)
        story.append(Spacer(1, 20))
        
        # Instructions
        story.append(Paragraph("Instructions", styles['Heading3']))
        instructions = [
            "1. Fill in client information in the KYC mapping fields",
            "2. Use templates for generating professional reports",
            "3. Maintain formatting consistency",
            "4. Include all required compliance information",
            "5. Export in required formats (Excel, PDF, Word)"
        ]
        
        for instruction in instructions:
            story.append(Paragraph(instruction, styles['Normal']))
        
        doc.build(story)
        logger.info(f"KYC-mapped PDF templates created successfully: {file_path}")
        
    except Exception as e:
        logger.error(f"Error creating KYC-mapped PDF templates: {str(e)}")
        raise

def create_kyc_mapped_word_templates(file_path, user):
    """Create Word templates with KYC mapping"""
    try:
        # For this implementation, we'll create a simple text-based template
        # In a production environment, you would use python-docx for proper Word documents
        
        content = f"""
FINANCIAL REPORT TEMPLATES
KYC-Mapped Professional Templates

CLIENT INFORMATION TEMPLATE
============================

Client Name: {user.get_full_name() if user else '[Client Name]'}
Company Name: [Company Name]
Email: {user.email if user else '[Email Address]'}
Report Period: [Report Period]
Prepared By: {user.get_full_name() if user else '[Preparer Name]'}
Date Prepared: {datetime.now().strftime('%Y-%m-%d')}

AVAILABLE REPORT TEMPLATES
==========================

1. Journal Report Template
   - Detailed transaction entries
   - Date, Reference, Account, Description, Amounts

2. Ledger Report Template
   - Account-wise transaction summaries
   - Running balances

3. Trial Balance Template
   - Account balances verification
   - Debit and credit totals

4. Profit & Loss Statement Template
   - Revenue and expense analysis
   - Net profit calculation

5. Balance Sheet Template
   - Assets, liabilities, and equity
   - Financial position statement

6. Cash Flow Statement Template
   - Operating, investing, financing activities
   - Cash movement analysis

7. Shareholders' Equity Template
   - Equity changes tracking
   - Capital structure analysis

8. MIS Report Template
   - Financial ratios and KPIs
   - Management information analysis

USAGE INSTRUCTIONS
==================

1. Fill in all KYC client information fields
2. Replace placeholder values with actual data
3. Maintain professional formatting
4. Include compliance requirements
5. Export in required formats

FEATURES
========

• Automated client branding integration
• Professional presentation formatting
• Compliance structure included
• Multi-format export capability
• KYC mapping for personalization

© F-AI Accountant - Professional Financial Reporting System
"""
        
        # Save as text file (in production, use python-docx for proper Word format)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"KYC-mapped Word templates created successfully: {file_path}")
        
    except Exception as e:
        logger.error(f"Error creating KYC-mapped Word templates: {str(e)}")
        raise