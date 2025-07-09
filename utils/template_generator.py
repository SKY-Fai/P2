"""
Template Generator for AccuFin360
Creates downloadable Excel templates for different data types
"""

import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class TemplateGenerator:
    """Generate Excel templates for different accounting data types"""
    
    def __init__(self):
        self.templates = {
            'general': self.create_general_template,
            'invoice': self.create_invoice_template,
            'inventory': self.create_inventory_template,
            'gst': self.create_gst_template,
            'payroll': self.create_payroll_template,
            'cash_flow': self.create_cash_flow_template,
            'budget': self.create_budget_template,
            # New Automated Accounting Templates
            'purchase': self.create_purchase_template,
            'sales': self.create_sales_template,
            'income': self.create_income_template,
            'expense': self.create_expense_template,
            'credit_note': self.create_credit_note_template,
            'debit_note': self.create_debit_note_template,
            'merged': self.create_merged_template,
            'combined_all': self.create_combined_all_template,
            'comprehensive_merged': self.create_comprehensive_merged_template,
            'payroll': self.generate_payroll_template,
            'bank_transfer': self.generate_bank_transfer_template,
            'asset_purchase': self.generate_asset_purchase_template,
            'others': self.generate_others_template
        }
    
    def generate_template(self, template_type: str, file_path: str) -> bool:
        """Generate template file"""
        try:
            if template_type not in self.templates:
                return False
            
            self.templates[template_type](file_path)
            return True
        except Exception as e:
            print(f"Error generating template: {str(e)}")
            return False
    
    def create_purchase_template(self) -> Workbook:
        """Create purchase transactions template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Template"
        
        # Headers with styling
        headers = [
            'Date', 'Vendor Name', 'Invoice Number', 'Amount', 
            'Tax Amount', 'Description', 'Payment Terms', 'Account Code'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'ABC Suppliers Ltd', 'INV-001', 10000.00, 1800.00, 'Office supplies purchase', '30 days', '5010'],
            ['2024-01-16', 'XYZ Services', 'INV-002', 5000.00, 900.00, 'Maintenance services', '15 days', '5160'],
            ['2024-01-17', 'Tech Solutions', 'INV-003', 25000.00, 4500.00, 'Software license', 'Net 30', '5140']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        return wb
    
    def create_sales_template(self) -> Workbook:
        """Create sales transactions template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Template"
        
        headers = [
            'Date', 'Customer Name', 'Invoice Number', 'Amount', 
            'Tax Amount', 'Description', 'Payment Due Date', 'Account Code'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Customer A Ltd', 'SAL-001', 15000.00, 2700.00, 'Product sales', '2024-02-15', '4010'],
            ['2024-01-16', 'Customer B Corp', 'SAL-002', 8000.00, 1440.00, 'Service revenue', '2024-02-01', '4020'],
            ['2024-01-17', 'Customer C Inc', 'SAL-003', 12000.00, 2160.00, 'Consulting services', '2024-02-10', '4020']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._adjust_column_widths(ws)
        return wb
    
    def create_income_template(self) -> Workbook:
        """Create income transactions template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Template"
        
        headers = [
            'Date', 'Source', 'Amount', 'Category', 
            'Description', 'Reference Number', 'Bank Account'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Interest Income', 500.00, 'interest', 'Bank interest earned', 'INT-001', 'SBI-12345'],
            ['2024-01-16', 'Rental Income', 15000.00, 'rental', 'Office space rental', 'RENT-001', 'HDFC-67890'],
            ['2024-01-17', 'Other Income', 2000.00, 'other', 'Miscellaneous income', 'MISC-001', 'ICICI-54321']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._adjust_column_widths(ws)
        return wb
    
    def create_expense_template(self) -> Workbook:
        """Create expense transactions template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Template"
        
        headers = [
            'Date', 'Vendor', 'Amount', 'Category', 
            'Description', 'Payment Method', 'Receipt Number'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Utility Company', 3000.00, 'utilities', 'Electricity bill', 'Bank Transfer', 'UTIL-001'],
            ['2024-01-16', 'Office Rent', 25000.00, 'rent', 'Monthly office rent', 'Cheque', 'RENT-001'],
            ['2024-01-17', 'Staff Salary', 150000.00, 'salary', 'Monthly salary payment', 'Bank Transfer', 'SAL-001']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._adjust_column_widths(ws)
        return wb
    
    def create_credit_note_template(self) -> Workbook:
        """Create credit note template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Note Template"
        
        headers = [
            'Date', 'Party Name', 'Original Invoice', 'Amount', 
            'Reason', 'Credit Note Number', 'Approval Status'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Customer A Ltd', 'INV-001', 2000.00, 'Product return', 'CN-001', 'Approved'],
            ['2024-01-16', 'Customer B Corp', 'INV-002', 1500.00, 'Billing error', 'CN-002', 'Pending'],
            ['2024-01-17', 'Customer C Inc', 'INV-003', 800.00, 'Discount adjustment', 'CN-003', 'Approved']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._adjust_column_widths(ws)
        return wb
    
    def create_debit_note_template(self) -> Workbook:
        """Create debit note template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Debit Note Template"
        
        headers = [
            'Date', 'Party Name', 'Original Invoice', 'Amount', 
            'Reason', 'Debit Note Number', 'Approval Status'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Supplier A Ltd', 'PUR-001', 1000.00, 'Additional charges', 'DN-001', 'Approved'],
            ['2024-01-16', 'Customer B Corp', 'SAL-002', 500.00, 'Late payment penalty', 'DN-002', 'Pending'],
            ['2024-01-17', 'Vendor C Inc', 'PUR-003', 750.00, 'Transport charges', 'DN-003', 'Approved']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._adjust_column_widths(ws)
        return wb
    
    def create_merged_template(self) -> Workbook:
        """Create comprehensive merged template for all transaction types"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Accounting Template"
        
        headers = [
            'Date', 'Transaction Type', 'Party Name', 'Amount', 
            'Description', 'Account Code', 'Reference', 'Tax Amount', 'Category'
        ]
        
        self._style_headers(ws, len(headers))
        
        # Sample data covering all transaction types
        sample_data = [
            ['2024-01-15', 'purchase', 'ABC Suppliers Ltd', 10000.00, 'Office supplies purchase', '5010', 'INV-001', 1800.00, 'supplies'],
            ['2024-01-16', 'sales', 'Customer A Ltd', 15000.00, 'Product sales', '4010', 'SAL-001', 2700.00, 'products'],
            ['2024-01-17', 'income', 'Interest Income', 500.00, 'Bank interest earned', '4110', 'INT-001', 0.00, 'interest'],
            ['2024-01-18', 'expense', 'Utility Company', 3000.00, 'Electricity bill', '5130', 'UTIL-001', 540.00, 'utilities'],
            ['2024-01-19', 'credit_note', 'Customer B Corp', 2000.00, 'Product return', '4030', 'CN-001', 360.00, 'returns'],
            ['2024-01-20', 'debit_note', 'Supplier C Ltd', 1000.00, 'Additional charges', '2010', 'DN-001', 180.00, 'adjustments']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add transaction type validation
        from openpyxl.worksheet.datavalidation import DataValidation
        
        transaction_types = '"purchase,sales,income,expense,credit_note,debit_note,payment,receipt,transfer,adjustment"'
        dv = DataValidation(type="list", formula1=transaction_types, allow_blank=False)
        dv.error = 'Please select a valid transaction type'
        dv.errorTitle = 'Invalid Transaction Type'
        ws.add_data_validation(dv)
        dv.add('B2:B1000')  # Apply to transaction type column
        
        self._adjust_column_widths(ws)
        
        # Add instructions sheet
        instructions_ws = wb.create_sheet(title="Instructions")
        instructions = [
            ['F-AI Accountant - Merged Accounting Template Instructions'],
            [''],
            ['Transaction Types:'],
            ['• purchase: Purchase transactions from vendors'],
            ['• sales: Sales transactions to customers'],
            ['• income: Income from various sources'],
            ['• expense: Business expense transactions'],
            ['• credit_note: Credit notes for returns/adjustments'],
            ['• debit_note: Debit notes for additional charges'],
            ['• payment: Payment transactions'],
            ['• receipt: Receipt transactions'],
            ['• transfer: Transfer between accounts'],
            ['• adjustment: Manual adjustments'],
            [''],
            ['Required Fields:'],
            ['• Date: Transaction date (YYYY-MM-DD format)'],
            ['• Transaction Type: Select from dropdown'],
            ['• Party Name: Vendor/Customer/Source name'],
            ['• Amount: Transaction amount (numeric)'],
            ['• Description: Transaction description'],
            ['• Account Code: Chart of accounts code'],
            ['• Reference: Invoice/Reference number'],
            [''],
            ['Optional Fields:'],
            ['• Tax Amount: Tax component if applicable'],
            ['• Category: Transaction category for classification'],
            [''],
            ['Processing:'],
            ['1. Fill in all required data'],
            ['2. Upload template to F-AI Accountant'],
            ['3. System will validate and process automatically'],
            ['4. Financial reports will be generated'],
            ['5. Journal entries will be posted']
        ]
        
        for row_idx, instruction in enumerate(instructions, start=1):
            instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
        
        return wb
    
    def create_general_template(self, file_path: str):
        """Create general accounting template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "General Accounting"
        
        # Headers
        headers = [
            'Date', 'Description', 'Account', 'Debit', 'Credit', 
            'Reference', 'Category', 'Notes'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['2025-01-01', 'Office rent payment', 'Rent Expense', 2500.00, '', 'CHQ001', 'Operating', 'Monthly rent'],
            ['2025-01-01', 'Office rent payment', 'Cash', '', 2500.00, 'CHQ001', 'Operating', 'Monthly rent'],
            ['2025-01-02', 'Sales revenue', 'Cash', 5000.00, '', 'INV001', 'Revenue', 'Product sales'],
            ['2025-01-02', 'Sales revenue', 'Sales Revenue', '', 5000.00, 'INV001', 'Revenue', 'Product sales'],
            ['2025-01-03', 'Equipment purchase', 'Equipment', 3000.00, '', 'PO001', 'Capital', 'Computer equipment'],
            ['2025-01-03', 'Equipment purchase', 'Accounts Payable', '', 3000.00, 'PO001', 'Capital', 'Computer equipment']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Add instructions sheet
        instructions = wb.create_sheet("Instructions")
        instructions['A1'] = "AccuFin360 - General Accounting Template"
        instructions['A1'].font = Font(bold=True, size=16)
        
        instruction_text = [
            "",
            "How to use this template:",
            "1. Fill in the Date column with transaction dates (YYYY-MM-DD format)",
            "2. Provide a clear Description for each transaction",
            "3. Specify the Account name or code affected",
            "4. Enter amounts in either Debit or Credit columns (not both)",
            "5. Add Reference numbers for tracking (invoice numbers, check numbers, etc.)",
            "6. Categorize transactions for better reporting",
            "7. Add Notes for additional context",
            "",
            "Important Notes:",
            "- Each transaction must have equal debits and credits",
            "- Use consistent account names throughout",
            "- Dates should be in YYYY-MM-DD format",
            "- Amounts should be positive numbers",
            "- Reference numbers help with audit trails"
        ]
        
        for i, text in enumerate(instruction_text, 2):
            instructions[f'A{i}'] = text
        
        wb.save(file_path)
    
    def create_invoice_template(self, file_path: str):
        """Create invoice template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"
        
        headers = [
            'Invoice Number', 'Customer Name', 'Customer Email', 'Invoice Date',
            'Due Date', 'Item Description', 'Quantity', 'Unit Price', 
            'Total Price', 'Tax Rate', 'Tax Amount', 'Invoice Total', 'Status'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['INV001', 'Acme Corp', 'billing@acme.com', '2025-01-01', '2025-01-31', 
             'Consulting Services', 10, 150.00, 1500.00, 0.18, 270.00, 1770.00, 'sent'],
            ['INV002', 'Tech Solutions', 'accounts@techsol.com', '2025-01-02', '2025-02-01', 
             'Software License', 5, 200.00, 1000.00, 0.18, 180.00, 1180.00, 'draft'],
            ['INV003', 'Global Industries', 'finance@global.com', '2025-01-03', '2025-02-02', 
             'Training Program', 20, 75.00, 1500.00, 0.18, 270.00, 1770.00, 'paid']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def create_inventory_template(self, file_path: str):
        """Create inventory template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Data"
        
        headers = [
            'Item Code', 'Item Name', 'Category', 'Description', 'Unit of Measure',
            'Current Stock', 'Reorder Level', 'Unit Cost', 'Selling Price', 
            'Supplier', 'Location', 'Last Updated'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='fd7e14', end_color='fd7e14', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['IT001', 'Laptop Computer', 'Electronics', 'Business laptop with 16GB RAM', 'Each', 
             25, 10, 800.00, 1200.00, 'Tech Supplier Inc', 'Warehouse A', '2025-01-01'],
            ['OF001', 'Office Chair', 'Furniture', 'Ergonomic office chair', 'Each', 
             15, 5, 150.00, 250.00, 'Office Furniture Ltd', 'Warehouse B', '2025-01-01'],
            ['ST001', 'Printer Paper', 'Supplies', 'A4 white paper 80gsm', 'Ream', 
             100, 20, 5.00, 8.00, 'Paper Supply Co', 'Storage Room', '2025-01-01']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def create_gst_template(self, file_path: str):
        """Create GST template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "GST Records"
        
        headers = [
            'GSTIN', 'Invoice Number', 'Invoice Date', 'Customer/Supplier Name',
            'Taxable Amount', 'CGST Rate', 'CGST Amount', 'SGST Rate', 'SGST Amount',
            'IGST Rate', 'IGST Amount', 'Total Tax', 'Total Amount', 'Return Period'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['27ABCDE1234F1Z5', 'INV001', '2025-01-01', 'ABC Customer Ltd', 
             10000.00, 9.0, 900.00, 9.0, 900.00, 0.0, 0.00, 1800.00, 11800.00, '01-2025'],
            ['27ABCDE1234F1Z5', 'INV002', '2025-01-02', 'XYZ Services Pvt Ltd', 
             15000.00, 0.0, 0.00, 0.0, 0.00, 18.0, 2700.00, 2700.00, 17700.00, '01-2025'],
            ['27ABCDE1234F1Z5', 'INV003', '2025-01-03', 'PQR Industries', 
             8000.00, 6.0, 480.00, 6.0, 480.00, 0.0, 0.00, 960.00, 8960.00, '01-2025']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def create_payroll_template(self, file_path: str):
        """Create payroll template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Data"
        
        headers = [
            'Employee ID', 'Employee Name', 'Department', 'Designation', 'Pay Period',
            'Basic Salary', 'Allowances', 'Overtime', 'Gross Salary', 'PF Deduction',
            'Tax Deduction', 'Other Deductions', 'Net Salary', 'Payment Date'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['EMP001', 'John Smith', 'Accounting', 'Senior Accountant', '2025-01', 
             50000.00, 10000.00, 5000.00, 65000.00, 6000.00, 8000.00, 1000.00, 50000.00, '2025-01-31'],
            ['EMP002', 'Jane Doe', 'Finance', 'Financial Analyst', '2025-01', 
             45000.00, 8000.00, 3000.00, 56000.00, 5400.00, 6500.00, 500.00, 43600.00, '2025-01-31'],
            ['EMP003', 'Bob Johnson', 'IT', 'System Administrator', '2025-01', 
             55000.00, 12000.00, 4000.00, 71000.00, 6600.00, 9000.00, 800.00, 54600.00, '2025-01-31']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def create_cash_flow_template(self, file_path: str):
        """Create cash flow template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Data"
        
        headers = [
            'Date', 'Description', 'Category', 'Cash Inflow', 'Cash Outflow',
            'Net Cash Flow', 'Balance', 'Source/Destination', 'Reference'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0dcaf0', end_color='0dcaf0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['2025-01-01', 'Customer payment received', 'Operating', 25000.00, 0.00, 
             25000.00, 25000.00, 'ABC Customer', 'RCP001'],
            ['2025-01-02', 'Office rent payment', 'Operating', 0.00, 5000.00, 
             -5000.00, 20000.00, 'Property Owner', 'CHQ001'],
            ['2025-01-03', 'Equipment purchase', 'Investing', 0.00, 15000.00, 
             -15000.00, 5000.00, 'Tech Equipment Ltd', 'PO001']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def create_budget_template(self, file_path: str):
        """Create budget template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Data"
        
        headers = [
            'Account Category', 'Account Name', 'Q1 Budget', 'Q2 Budget', 'Q3 Budget', 
            'Q4 Budget', 'Annual Budget', 'Q1 Actual', 'Q2 Actual', 'Q3 Actual', 
            'Q4 Actual', 'Annual Actual', 'Variance'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='20c997', end_color='20c997', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['Revenue', 'Sales Revenue', 100000, 110000, 120000, 130000, 460000, 
             95000, 105000, 115000, 125000, 440000, -20000],
            ['Expenses', 'Office Rent', 10000, 10000, 10000, 10000, 40000, 
             10000, 10000, 10000, 10000, 40000, 0],
            ['Expenses', 'Marketing', 15000, 16000, 17000, 18000, 66000, 
             12000, 14000, 16000, 19000, 61000, -5000]
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
    
    def generate_payroll_template(self, file_path: str):
        """Generate payroll template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Template"
        
        # Headers
        headers = [
            'Date', 'Employee Name', 'Employee ID', 'Department',
            'Basic Salary', 'Allowances', 'Overtime', 'Deductions',
            'Gross Salary', 'Tax Deducted', 'Net Salary', 'Account Code',
            'Payment Method', 'Reference'
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        sample_data = [
            ['2024-01-31', 'John Doe', 'EMP001', 'Finance', 50000, 5000, 2000, 1000, 56000, 8400, 47600, '5010', 'Bank Transfer', 'PAY2024001'],
            ['2024-01-31', 'Jane Smith', 'EMP002', 'HR', 45000, 4500, 1000, 500, 50000, 7500, 42500, '5010', 'Bank Transfer', 'PAY2024002']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._style_headers(ws, len(headers))
        self._adjust_column_widths(ws)
        
        wb.save(file_path)
        return file_path
    
    def generate_bank_transfer_template(self, file_path: str):
        """Generate bank transfer template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Transfer Template"
        
        # Headers
        headers = [
            'Date', 'From Bank Account', 'To Bank Account', 'Amount',
            'Transfer Type', 'Reference Number', 'Description',
            'From Account Code', 'To Account Code', 'Currency'
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        sample_data = [
            ['2024-01-15', 'Current Account', 'Savings Account', 10000, 'Internal Transfer', 'TRF001', 'Fund transfer to savings', '1010', '1020', 'INR'],
            ['2024-01-20', 'Savings Account', 'Current Account', 5000, 'Internal Transfer', 'TRF002', 'Transfer for operations', '1020', '1010', 'INR']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._style_headers(ws, len(headers))
        self._adjust_column_widths(ws)
        
        wb.save(file_path)
        return file_path
    
    def generate_asset_purchase_template(self, file_path: str):
        """Generate asset purchase template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Asset Purchase Template"
        
        # Headers
        headers = [
            'Date', 'Asset Name', 'Asset Category', 'Vendor Name',
            'Purchase Amount', 'Tax Amount', 'Total Amount', 'Asset Code',
            'Depreciation Method', 'Useful Life (Years)', 'Account Code',
            'Invoice Number', 'Description'
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        sample_data = [
            ['2024-01-10', 'Office Computer', 'IT Equipment', 'Tech Suppliers Ltd', 50000, 9000, 59000, 'ASSET001', 'Straight Line', 3, '1200', 'INV001', 'Dell Desktop Computer'],
            ['2024-01-15', 'Office Furniture', 'Furniture', 'Office World', 25000, 4500, 29500, 'ASSET002', 'Straight Line', 5, '1210', 'INV002', 'Executive Office Desk']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._style_headers(ws, len(headers))
        self._adjust_column_widths(ws)
        
        wb.save(file_path)
        return file_path
    
    def generate_others_template(self, file_path: str):
        """Generate miscellaneous transactions template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Others Template"
        
        # Headers
        headers = [
            'Date', 'Transaction Type', 'Party Name', 'Amount',
            'Description', 'Account Code', 'Reference Number',
            'Category', 'Department', 'Project Code'
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        sample_data = [
            ['2024-01-12', 'Miscellaneous Income', 'Interest Income', 1500, 'Bank interest received', '4100', 'MISC001', 'Financial Income', 'Finance', 'PROJ001'],
            ['2024-01-18', 'Adjustment', 'Rounding Adjustment', 5, 'Rounding difference adjustment', '6999', 'ADJ001', 'Adjustment', 'Accounts', 'PROJ002']
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        self._style_headers(ws, len(headers))
        self._adjust_column_widths(ws)
        
        wb.save(file_path)
        return file_path

    def get_available_templates(self) -> dict:
        """Get list of available templates"""
        return {
            'general': 'General Accounting Template',
            'invoice': 'Invoice Data Template',
            'inventory': 'Inventory Management Template',
            'gst': 'GST Records Template',
            'payroll': 'Payroll Data Template',
            'cash_flow': 'Cash Flow Template',
            'budget': 'Budget Planning Template',
            'bank_transfer': 'Bank Transfer Template',
            'asset_purchase': 'Asset Purchase Template',
            'others': 'Miscellaneous Transactions Template',
            'purchase': 'Purchase Transactions Template',
            'sales': 'Sales Transactions Template',
            'income': 'Income Transactions Template',
            'expense': 'Expense Transactions Template',
            'credit_note': 'Credit Note Template',
            'debit_note': 'Debit Note Template',
            'combined_all': 'Combined All Templates (Multi-Sheet)',
            'comprehensive_merged': 'Comprehensive Merged Template (Single Sheet)'
        }
    
    def _style_headers(self, ws, num_headers):
        """Style header row with professional formatting"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Define header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, num_headers + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for better readability"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_combined_all_template(self, file_path: str):
        """Create standardized combined template with all transaction types in one Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define professional styling
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data styling
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        amount_alignment = Alignment(horizontal='right', vertical='center')
        
        # Professional borders
        thick_border = Border(
            left=Side(style='thick', color='1F4E79'),
            right=Side(style='thick', color='1F4E79'),
            top=Side(style='thick', color='1F4E79'),
            bottom=Side(style='thick', color='1F4E79')
        )
        
        thin_border = Border(
            left=Side(style='thin', color='666666'),
            right=Side(style='thin', color='666666'),
            top=Side(style='thin', color='666666'),
            bottom=Side(style='thin', color='666666')
        )
        
        # Standardized template structure - all sheets follow consistent 10-column format
        templates_data = {
            'Purchase': {
                'headers': ['Date*', 'Party Name*', 'Document Number', 'Description*', 'Base Amount*', 'Tax Amount', 'Total Amount*', 'Payment Method', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'ABC Suppliers Ltd', 'INV-001', 'Office Supplies Purchase', '10000.00', '1800.00', '11800.00', 'Bank Transfer', 'PUR001', 'Monthly office supplies'],
                    ['2024-01-16', 'XYZ Services Pvt Ltd', 'INV-002', 'Consulting Services', '25000.00', '4500.00', '29500.00', 'Cheque', 'PUR002', 'Management consulting'],
                    ['2024-01-17', 'Tech Solutions Inc', 'INV-003', 'Software License', '50000.00', '9000.00', '59000.00', 'Online Payment', 'PUR003', 'Annual software subscription']
                ],
                'validations': ['Date*', 'Payment Method']
            },
            'Sales': {
                'headers': ['Date*', 'Party Name*', 'Document Number', 'Description*', 'Base Amount*', 'Tax Amount', 'Total Amount*', 'Payment Method', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'John Smith Enterprises', 'SALE-001', 'Product A Sales', '15000.00', '2700.00', '17700.00', 'Bank Transfer', 'SAL001', 'Bulk order discount applied'],
                    ['2024-01-16', 'Jane Doe Corporation', 'SALE-002', 'Service B Provision', '35000.00', '6300.00', '41300.00', 'Cheque', 'SAL002', 'Premium service package'],
                    ['2024-01-17', 'ABC Corp Limited', 'SALE-003', 'Product C Sales', '25000.00', '4500.00', '29500.00', 'Online Payment', 'SAL003', 'Regular customer order']
                ],
                'validations': ['Date*', 'Payment Method']
            },
            'Income': {
                'headers': ['Date*', 'Income Source*', 'Document Number', 'Description*', 'Base Amount*', 'Tax Amount', 'Net Amount*', 'Payment Method', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'Interest Income', 'INT-001', 'Bank Interest Received', '5000.00', '500.00', '4500.00', 'Bank Credit', 'INC001', 'Quarterly interest'],
                    ['2024-01-16', 'Rental Income', 'RENT-001', 'Property Rent Received', '100000.00', '10000.00', '90000.00', 'Bank Transfer', 'INC002', 'Monthly office rent'],
                    ['2024-01-17', 'Dividend Income', 'DIV-001', 'Share Dividend Received', '20000.00', '2000.00', '18000.00', 'Bank Credit', 'INC003', 'Annual dividend payment']
                ],
                'validations': ['Date*', 'Payment Method']
            },
            'Expense': {
                'headers': ['Date*', 'Expense Category*', 'Document Number', 'Description*', 'Base Amount*', 'Tax Amount', 'Total Amount*', 'Payment Method', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'Office Rent', 'RENT-001', 'Monthly Office Rent', '150000.00', '27000.00', '177000.00', 'Bank Transfer', 'EXP001', 'Head office monthly rent'],
                    ['2024-01-16', 'Utilities', 'ELEC-001', 'Electricity Bill Payment', '30000.00', '5400.00', '35400.00', 'Online Payment', 'EXP002', 'Monthly electricity charges'],
                    ['2024-01-17', 'Travel & Transport', 'TRAVEL-001', 'Business Travel Expenses', '50000.00', '9000.00', '59000.00', 'Credit Card', 'EXP003', 'Client meeting travel cost']
                ],
                'validations': ['Date*', 'Payment Method']
            },
            'Credit Note': {
                'headers': ['Date*', 'Customer Name*', 'Original Invoice', 'Credit Note No.*', 'Reason*', 'Base Amount*', 'Tax Amount', 'Total Credit*', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'John Smith Enterprises', 'SALE-001', 'CN-001', 'Product Return', '5000.00', '900.00', '5900.00', 'CN001', 'Defective product returned'],
                    ['2024-01-16', 'Jane Doe Corporation', 'SALE-002', 'CN-002', 'Service Adjustment', '10000.00', '1800.00', '11800.00', 'CN002', 'Service quality issue'],
                    ['2024-01-17', 'ABC Corp Limited', 'SALE-003', 'CN-003', 'Discount Applied', '2500.00', '450.00', '2950.00', 'CN003', 'Volume discount credit']
                ],
                'validations': ['Date*']
            },
            'Debit Note': {
                'headers': ['Date*', 'Vendor Name*', 'Original Invoice', 'Debit Note No.*', 'Reason*', 'Base Amount*', 'Tax Amount', 'Total Debit*', 'Reference', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'ABC Suppliers Ltd', 'INV-001', 'DN-001', 'Shortage in Supply', '2000.00', '360.00', '2360.00', 'DN001', 'Quantity shortage in delivery'],
                    ['2024-01-16', 'XYZ Services Pvt Ltd', 'INV-002', 'DN-002', 'Quality Issue', '5000.00', '900.00', '5900.00', 'DN002', 'Service quality below standard'],
                    ['2024-01-17', 'Tech Solutions Inc', 'INV-003', 'DN-003', 'Additional Charges', '3000.00', '540.00', '3540.00', 'DN003', 'Installation charges not included']
                ],
                'validations': ['Date*']
            },
            'Payroll': {
                'headers': ['Pay Period*', 'Employee Name*', 'Employee ID', 'Department*', 'Basic Salary*', 'Allowances', 'Overtime', 'Gross Salary*', 'Deductions', 'Net Salary*'],
                'sample_data': [
                    ['2024-01', 'John Doe', 'EMP001', 'Finance', '500000.00', '50000.00', '20000.00', '570000.00', '140000.00', '430000.00'],
                    ['2024-01', 'Jane Smith', 'EMP002', 'Human Resources', '450000.00', '45000.00', '15000.00', '510000.00', '125000.00', '385000.00'],
                    ['2024-01', 'Mike Johnson', 'EMP003', 'Information Technology', '600000.00', '60000.00', '30000.00', '690000.00', '170000.00', '520000.00']
                ],
                'validations': ['Pay Period*', 'Department*']
            },
            'Bank Transfer': {
                'headers': ['Date*', 'From Account*', 'To Account*', 'Transfer Amount*', 'Transfer Type*', 'Reference Number', 'Description*', 'Bank Charges', 'Net Amount', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'Savings Account - 12345', 'Current Account - 67890', '250000.00', 'Internal Transfer', 'TXN001', 'Working capital transfer', '100.00', '249900.00', 'Monthly fund allocation'],
                    ['2024-01-16', 'Current Account - 67890', 'FD Account - 11111', '500000.00', 'Internal Transfer', 'TXN002', 'Fixed deposit investment', '150.00', '499850.00', 'Quarterly FD investment'],
                    ['2024-01-17', 'Current Account - 67890', 'Vendor Account - EXT123', '100000.00', 'External Transfer', 'TXN003', 'Vendor payment transfer', '500.00', '99500.00', 'Monthly vendor settlement']
                ],
                'validations': ['Date*', 'Transfer Type*']
            },
            'Asset Purchase': {
                'headers': ['Purchase Date*', 'Asset Name*', 'Asset Category*', 'Vendor Name*', 'Purchase Amount*', 'Depreciation Method', 'Useful Life (Years)', 'Reference', 'Location', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'Dell Precision Workstation', 'Computer Equipment', 'Dell Technologies India', '750000.00', 'Straight Line Method', '3', 'ASSET001', 'Head Office', 'High-end workstation for design team'],
                    ['2024-01-16', 'Herman Miller Ergonomic Chairs', 'Furniture & Fixtures', 'Urban Ladder Solutions', '125000.00', 'Straight Line Method', '5', 'ASSET002', 'Head Office', 'Ergonomic chairs for employee comfort'],
                    ['2024-01-17', 'Toyota Innova Crysta', 'Transportation Vehicle', 'Toyota Motors India', '2500000.00', 'Written Down Value Method', '8', 'ASSET003', 'Corporate Fleet', 'Executive transport vehicle']
                ],
                'validations': ['Purchase Date*', 'Asset Category*']
            },
            'Others': {
                'headers': ['Date*', 'Transaction Type*', 'Description*', 'Debit Account*', 'Credit Account*', 'Amount*', 'Reference Number', 'Approval By', 'Document No', 'Notes'],
                'sample_data': [
                    ['2024-01-15', 'Journal Entry', 'Monthly Bank Charges', 'Bank Charges Expense', 'Primary Bank Account', '5000.00', 'JE001', 'Finance Manager', 'MISC-001', 'Standard monthly bank charges'],
                    ['2024-01-16', 'Adjustment Entry', 'Foreign Exchange Adjustment', 'Foreign Exchange Gain/Loss', 'Accounts Receivable', '20000.00', 'ADJ001', 'CFO', 'MISC-002', 'Quarterly FX rate adjustment'],
                    ['2024-01-17', 'Provision Entry', 'Bad Debt Provision', 'Bad Debt Expense', 'Provision for Doubtful Debts', '100000.00', 'PROV001', 'Finance Head', 'MISC-003', 'Annual bad debt provision']
                ],
                'validations': ['Date*', 'Transaction Type*']
            }
        }
        
        # Create each sheet
        for sheet_name, data in templates_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Add headers
            for col, header in enumerate(data['headers'], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Add sample data
            for row_idx, row_data in enumerate(data['sample_data'], 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create Instructions sheet
        instructions_ws = wb.create_sheet(title="Instructions", index=0)
        
        instructions = [
            ["F-AI ACCOUNTANT - STANDARDIZED COMBINED TRANSACTION TEMPLATES"],
            [""],
            ["OVERVIEW:"],
            ["This is a comprehensive Excel template package containing all 10 transaction types"],
            ["standardized for automated accounting processing with AI intelligence."],
            [""],
            ["STANDARDIZED FEATURES:"],
            ["✓ Consistent 10-column structure across all transaction types"],
            ["✓ Mandatory fields marked with asterisk (*)"],
            ["✓ Professional formatting and validation"],
            ["✓ Sample data for guidance"],
            ["✓ Compatible with F-AI Accounting Engine"],
            [""],
            ["INSTRUCTIONS FOR USE:"],
            [""],
            ["1. PREPARATION:"],
            ["   • This file contains 10 transaction sheets + this instruction sheet"],
            ["   • Each sheet follows a standardized 10-column format"],
            ["   • Required fields are marked with asterisk (*) in headers"],
            [""],
            ["2. DATA ENTRY:"],
            ["   • Replace sample data with your actual transactions"],
            ["   • Keep header row (Row 1) unchanged"],
            ["   • Fill all mandatory fields (marked with *)"],
            ["   • Date format: YYYY-MM-DD (e.g., 2024-01-15)"],
            ["   • Amount format: Numbers only, no currency symbols (e.g., 10000.00)"],
            [""],
            ["3. QUALITY CHECKS:"],
            ["   • Ensure consistent party/account naming"],
            ["   • Verify amount calculations"],
            ["   • Use proper date formats"],
            ["   • Fill reference numbers for tracking"],
            [""],
            ["4. UPLOAD PROCESS:"],
            ["   • Save file after completing data entry"],
            ["   • Upload to F-AI Accounting > AI Accounting module"],
            ["   • Select 'Combined All Templates' as template type"],
            ["   • Wait for AI processing and report generation"],
            [""],
            ["TRANSACTION SHEET DETAILS:"],
            [""],
            ["📋 PURCHASE: Vendor purchases, bills, and business expenses"],
            ["   Key Fields: Date*, Party Name*, Description*, Base Amount*, Total Amount*"],
            [""],
            ["📋 SALES: Customer sales, invoices, and revenue transactions"],
            ["   Key Fields: Date*, Party Name*, Description*, Base Amount*, Total Amount*"],
            [""],
            ["📋 INCOME: Other income sources (interest, rent, dividends)"],
            ["   Key Fields: Date*, Income Source*, Description*, Base Amount*, Net Amount*"],
            [""],
            ["📋 EXPENSE: Operational expenses and business costs"],
            ["   Key Fields: Date*, Expense Category*, Description*, Base Amount*, Total Amount*"],
            [""],
            ["📋 CREDIT NOTE: Sales returns, refunds, and customer adjustments"],
            ["   Key Fields: Date*, Customer Name*, Credit Note No.*, Reason*, Total Credit*"],
            [""],
            ["📋 DEBIT NOTE: Purchase returns and vendor adjustments"],
            ["   Key Fields: Date*, Vendor Name*, Debit Note No.*, Reason*, Total Debit*"],
            [""],
            ["📋 PAYROLL: Employee salary, wages, and compensation"],
            ["   Key Fields: Pay Period*, Employee Name*, Department*, Basic Salary*, Net Salary*"],
            [""],
            ["📋 BANK TRANSFER: Internal and external fund transfers"],
            ["   Key Fields: Date*, From Account*, To Account*, Transfer Amount*, Transfer Type*"],
            [""],
            ["📋 ASSET PURCHASE: Fixed asset acquisitions and investments"],
            ["   Key Fields: Purchase Date*, Asset Name*, Asset Category*, Vendor Name*, Purchase Amount*"],
            [""],
            ["📋 OTHERS: Miscellaneous transactions and journal entries"],
            ["   Key Fields: Date*, Transaction Type*, Description*, Debit Account*, Credit Account*, Amount*"],
            [""],
            ["BEST PRACTICES:"],
            [""],
            ["• Use only sheets relevant to your business operations"],
            ["• Delete sample rows after understanding the format"],
            ["• Maintain consistent naming conventions throughout"],
            ["• Include detailed descriptions for better tracking"],
            ["• Use reference numbers for audit trail"],
            ["• Validate totals before uploading"],
            [""],
            ["SUPPORT:"],
            ["For technical assistance or questions about F-AI Accounting,"],
            ["please refer to the user manual or contact support."],
            [""],
            ["Generated by F-AI Accountant | Professional Accounting Solution"]
        ]
        
        # Style instructions
        title_font = Font(bold=True, size=14, color='FFFFFF')
        title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        section_font = Font(bold=True, size=12, color='2F4F4F')
        normal_font = Font(size=10)
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
            
            if row_idx == 1:  # Title
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = Alignment(horizontal='center')
            elif instruction[0].endswith(':') and instruction[0].isupper():  # Section headers
                cell.font = section_font
            else:  # Normal text
                cell.font = normal_font
        
        # Adjust column width for instructions
        instructions_ws.column_dimensions['A'].width = 80
        
        wb.save(file_path)

    def create_comprehensive_merged_template(self) -> Workbook:
        """Create a single comprehensive template with all categories organized by transaction type"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Complete Accounting Template"
        
        # Define comprehensive headers for all transaction categories
        headers = [
            # Core Transaction Fields
            'Transaction_Type',  # Purchase, Sales, Income, Expense, Credit_Note, Debit_Note, etc.
            'Category',         # General, Invoice, Inventory, GST, Payroll, etc.
            'Date',
            'Description',
            'Reference_Number',
            'Amount',
            'Currency',
            
            # Party Information
            'Party_Name',
            'Party_Type',       # Customer, Vendor, Employee, etc.
            'Party_Address',
            'Party_Phone',
            'Party_Email',
            'Party_GST_Number',
            'Party_PAN',
            
            # Account Details
            'Debit_Account',
            'Credit_Account',
            'Account_Code',
            'Cost_Center',
            'Project_Code',
            
            # Tax Information
            'Tax_Type',         # GST, VAT, etc.
            'Tax_Rate',
            'Tax_Amount',
            'CGST_Rate',
            'CGST_Amount',
            'SGST_Rate', 
            'SGST_Amount',
            'IGST_Rate',
            'IGST_Amount',
            'CESS_Rate',
            'CESS_Amount',
            
            # Item/Product Details
            'Item_Name',
            'Item_Code',
            'Item_Category',
            'Quantity',
            'Unit_Price',
            'Unit_of_Measure',
            'HSN_SAC_Code',
            
            # Invoice Specific
            'Invoice_Number',
            'Invoice_Date',
            'Due_Date',
            'Payment_Terms',
            'Discount_Percentage',
            'Discount_Amount',
            
            # Payment Details
            'Payment_Method',   # Cash, Bank, Card, etc.
            'Bank_Account',
            'Bank_Reference',
            'Payment_Status',   # Paid, Pending, Partial, etc.
            
            # Payroll Specific
            'Employee_ID',
            'Employee_Name',
            'Department',
            'Designation',
            'Basic_Salary',
            'Allowances',
            'Deductions',
            'Net_Salary',
            
            # Inventory Specific
            'Warehouse',
            'Batch_Number',
            'Serial_Number',
            'Expiry_Date',
            'Manufacturing_Date',
            
            # Additional Fields
            'Notes',
            'Approval_Status',
            'Created_By',
            'Modified_By',
            'Status'            # Active, Inactive, Draft, etc.
        ]
        
        # Apply header styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add sample data for each category
        sample_data = self._get_comprehensive_sample_data()
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                # Center align certain columns
                if col_idx in [1, 2, 3, 6, 7]:  # Transaction_Type, Category, Date, Amount, Currency
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add data validation and instructions sheet
        instructions_ws = wb.create_sheet(title="Instructions")
        self._add_template_instructions(instructions_ws)
        
        # Add dropdown validation sheet
        validation_ws = wb.create_sheet(title="Validation_Data")
        self._add_validation_data(validation_ws)
        
        return wb
    
    def _get_comprehensive_sample_data(self):
        """Generate comprehensive sample data for all categories"""
        return [
            # Purchase Transaction
            ['Purchase', 'General', '2024-01-15', 'Office Supplies Purchase', 'PO-001', 15000, 'INR',
             'ABC Suppliers Ltd', 'Vendor', '123 Business Park, Mumbai', '+91-9876543210', 'supplier@abc.com',
             '27AAAAA0000A1Z5', 'AAAAA0000A', 'Office Expenses', 'Accounts Payable', 'OE001', 'CC001', 'PROJ001',
             'GST', 18, 2700, 9, 1350, 9, 1350, 0, 0, 0, 0,
             'A4 Paper, Pens, Files', 'SUP001', 'Stationery', 100, 150, 'Pieces', '48191000',
             'INV-ABC-001', '2024-01-15', '2024-02-14', '30 Days', 0, 0,
             'Bank Transfer', 'SBI-001', 'TXN123456', 'Pending',
             '', '', '', '', 0, 0, 0, 0,
             'Main Warehouse', 'BATCH001', '', '', '',
             'Monthly office supplies purchase', 'Approved', 'Admin', 'Admin', 'Active'],
            
            # Sales Transaction
            ['Sales', 'Invoice', '2024-01-16', 'Product Sale to Customer', 'SO-001', 25000, 'INR',
             'XYZ Enterprises', 'Customer', '456 Market Street, Delhi', '+91-9876543211', 'customer@xyz.com',
             '07BBBBB0000B1Z5', 'BBBBB0000B', 'Sales Revenue', 'Accounts Receivable', 'SR001', 'CC002', 'PROJ002',
             'GST', 18, 4500, 9, 2250, 9, 2250, 0, 0, 0, 0,
             'Software License', 'PROD001', 'Software', 5, 5000, 'Licenses', '99831990',
             'INV-001', '2024-01-16', '2024-02-15', '30 Days', 10, 2500,
             'Bank Transfer', 'HDFC-001', 'TXN789123', 'Paid',
             '', '', '', '', 0, 0, 0, 0,
             'Digital Delivery', '', '', '', '',
             'Software license sale with discount', 'Approved', 'Sales', 'Sales', 'Active'],
            
            # Payroll Transaction
            ['Expense', 'Payroll', '2024-01-31', 'January Salary Payment', 'PAY-001', 50000, 'INR',
             'John Doe', 'Employee', 'Mumbai', '+91-9876543212', 'john@company.com',
             '', 'CCCCC0000C', 'Salary Expense', 'Bank Account', 'SE001', 'HR001', 'DEPT001',
             '', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
             '', '', '', 1, 50000, 'Month', '',
             '', '', '', '', 0, 0,
             'Bank Transfer', 'ICICI-001', 'SAL123456', 'Paid',
             'EMP001', 'John Doe', 'IT Department', 'Software Developer', 40000, 10000, 5000, 45000,
             '', '', '', '', '',
             'Monthly salary payment', 'Approved', 'HR', 'HR', 'Active'],
            
            # GST Transaction
            ['Purchase', 'GST', '2024-01-20', 'Raw Material Purchase', 'PO-002', 100000, 'INR',
             'Raw Material Suppliers', 'Vendor', '789 Industrial Area, Chennai', '+91-9876543213', 'rm@suppliers.com',
             '33DDDDD0000D1Z5', 'DDDDD0000D', 'Raw Materials', 'Accounts Payable', 'RM001', 'PROD001', 'MFG001',
             'GST', 18, 18000, 9, 9000, 9, 9000, 0, 0, 2, 2000,
             'Steel Sheets', 'RM001', 'Raw Material', 500, 200, 'KG', '72081000',
             'INV-RM-001', '2024-01-20', '2024-02-19', '30 Days', 0, 0,
             'Bank Transfer', 'AXIS-001', 'TXN456789', 'Pending',
             '', '', '', '', 0, 0, 0, 0,
             'Raw Material Warehouse', 'BATCH002', '', '2025-12-31', '2024-01-15',
             'High grade steel for manufacturing', 'Approved', 'Purchase', 'Purchase', 'Active'],
            
            # Inventory Transaction
            ['Income', 'Inventory', '2024-01-25', 'Inventory Sale', 'IS-001', 75000, 'INR',
             'Wholesale Customer', 'Customer', 'Bangalore', '+91-9876543214', 'wholesale@customer.com',
             '29EEEEE0000E1Z5', 'EEEEE0000E', 'Inventory Sales', 'Cash Account', 'IS001', 'INV001', 'SALE001',
             'GST', 18, 13500, 9, 6750, 9, 6750, 0, 0, 0, 0,
             'Finished Goods', 'FG001', 'Finished Product', 100, 750, 'Units', '84131900',
             'INV-WS-001', '2024-01-25', '2024-01-25', 'Immediate', 5, 3750,
             'Cash', 'CASH-001', 'CASH12345', 'Paid',
             '', '', '', '', 0, 0, 0, 0,
             'Finished Goods Warehouse', 'BATCH003', 'SN001-100', '', '2024-01-20',
             'Wholesale inventory sale', 'Approved', 'Inventory', 'Inventory', 'Active']
        ]
    
    def _add_template_instructions(self, ws):
        """Add detailed instructions for using the template"""
        instructions = [
            "COMPREHENSIVE ACCOUNTING TEMPLATE - INSTRUCTIONS",
            "",
            "1. OVERVIEW:",
            "   This template supports all major accounting transactions in a single file.",
            "   Each row represents one transaction with comprehensive details.",
            "",
            "2. MANDATORY FIELDS:",
            "   - Transaction_Type: Purchase, Sales, Income, Expense, Credit_Note, Debit_Note",
            "   - Category: General, Invoice, Inventory, GST, Payroll, etc.",
            "   - Date: Format YYYY-MM-DD",
            "   - Description: Clear description of transaction",
            "   - Amount: Transaction amount (without tax)",
            "",
            "3. TRANSACTION TYPES:",
            "   Purchase: Buying goods/services from vendors",
            "   Sales: Selling goods/services to customers", 
            "   Income: Revenue from various sources",
            "   Expense: Business expenses and costs",
            "   Credit_Note: Credits issued to customers",
            "   Debit_Note: Debits issued to vendors",
            "",
            "4. CATEGORIES:",
            "   General: Standard business transactions",
            "   Invoice: Formal billing transactions",
            "   Inventory: Stock/product related transactions",
            "   GST: Tax-focused transactions",
            "   Payroll: Employee salary and benefits",
            "",
            "5. TAX CALCULATIONS:",
            "   - Enter tax rates as percentages (e.g., 18 for 18%)",
            "   - System will calculate CGST/SGST for intra-state",
            "   - Use IGST for inter-state transactions",
            "",
            "6. PARTY INFORMATION:",
            "   - Include complete vendor/customer details",
            "   - GST numbers mandatory for B2B transactions",
            "   - PAN required for high-value transactions",
            "",
            "7. ACCOUNT MAPPING:",
            "   - Use standard chart of accounts codes",
            "   - Ensure proper debit/credit account assignment",
            "   - Include cost center and project codes if applicable",
            "",
            "8. DATA VALIDATION:",
            "   - Check 'Validation_Data' sheet for dropdown options",
            "   - Ensure data consistency across all fields",
            "   - Validate all calculations before submission",
            "",
            "9. BEST PRACTICES:",
            "   - Fill all relevant fields for complete records",
            "   - Use consistent naming conventions",
            "   - Include reference numbers for tracking",
            "   - Add detailed notes for complex transactions",
            "",
            "10. SUPPORT:",
            "    Contact your accounting team for assistance with:",
            "    - Account code mapping",
            "    - Tax rate determination", 
            "    - Approval workflows",
            "    - System integration issues"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row, column=1, value=instruction)
            if row == 1:
                cell.font = Font(bold=True, size=14, color="2C3E50")
            elif instruction.endswith(":"):
                cell.font = Font(bold=True, color="2C3E50")
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 80
    
    def _add_validation_data(self, ws):
        """Add validation data for dropdowns"""
        validation_data = {
            'Transaction_Types': ['Purchase', 'Sales', 'Income', 'Expense', 'Credit_Note', 'Debit_Note'],
            'Categories': ['General', 'Invoice', 'Inventory', 'GST', 'Payroll', 'Cash_Flow', 'Budget'],
            'Party_Types': ['Customer', 'Vendor', 'Employee', 'Bank', 'Government', 'Others'],
            'Payment_Methods': ['Cash', 'Bank Transfer', 'Credit Card', 'Debit Card', 'Cheque', 'UPI', 'Online'],
            'Payment_Status': ['Paid', 'Pending', 'Partial', 'Overdue', 'Cancelled'],
            'Currencies': ['INR', 'USD', 'EUR', 'GBP', 'JPY', 'SGD', 'AED'],
            'Tax_Types': ['GST', 'VAT', 'Service Tax', 'Excise', 'Customs', 'Income Tax', 'None'],
            'Status_Options': ['Active', 'Inactive', 'Draft', 'Approved', 'Rejected', 'Pending'],
            'Approval_Status': ['Pending', 'Approved', 'Rejected', 'Under Review']
        }
        
        col = 1
        for category, options in validation_data.items():
            # Add category header
            ws.cell(row=1, column=col, value=category).font = Font(bold=True)
            
            # Add options
            for row, option in enumerate(options, 2):
                ws.cell(row=row, column=col, value=option)
            
            # Auto-adjust column width
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
            col += 1